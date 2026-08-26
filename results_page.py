import re
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

from PySide6.QtCore import Qt, QTimer
from PySide6.QtWidgets import (
    QAbstractItemView,
    QComboBox,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QStyledItemDelegate,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
    QFileDialog,   # <-- ensure this is here
)

import excel_utils
from class_utils import get_classes
from db_utils import get_cursor, fetch_all, fetch_one, get_exam_context
from event_bus import EventBus
from system_state import SystemState
from ui_helpers import show_error, show_info
import combo_loaders


def _subject_name_matches(candidate, target):
    """Match subject names while ignoring display suffixes and case."""
    if candidate is None or target is None:
        return False

    def normalize(value):
        text = str(value).strip().lower()
        text = re.sub(r"\s+", " ", text)
        text = re.sub(r"\s*\(\d+%\)\s*$", "", text)
        text = re.sub(r"\s*\([^)]+\)$", "", text)
        text = re.sub(r"\s*\[[^\]]+\]$", "", text)
        return text

    return normalize(candidate) == normalize(target)


class MarksDelegate(QStyledItemDelegate):

    def createEditor(self, parent, option, index):
        editor = QLineEdit(parent)
        editor.setObjectName("MarksEditor")
        editor.setAlignment(Qt.AlignmentFlag.AlignCenter)
        editor.setPlaceholderText("0 - 100")
        editor.setFrame(True)

        editor.textChanged.connect(
            lambda text, field=editor, model=index.model(), cell=index:
            self._handle_text_changed(field, model, cell, text)
        )

        return editor

    def setEditorData(self, editor, index):
        editor.setText(str(index.data() or ""))
        self._set_validation_style(editor, editor.text())

    def setModelData(self, editor, model, index):
        model.setData(
            index,
            editor.text().strip(),
            Qt.ItemDataRole.EditRole,
        )

    def updateEditorGeometry(self, editor, option, index):
        editor.setGeometry(option.rect.adjusted(4, 3, -4, -3))

    def _handle_text_changed(self, editor, model, index, text):
        self._set_validation_style(editor, text)
        model.setData(
            index,
            text,
            Qt.ItemDataRole.EditRole,
        )

    @staticmethod
    def _set_validation_style(editor, text):
        value = text.strip()
        valid = not value or (
            value.isdigit()
            and 0 <= int(value) <= 100
        )

        if valid:
            editor.setProperty("variant", "accent")
        else:
            editor.setProperty("variant", "danger")


class ResultsPage(QWidget):

    def __init__(self):
        super().__init__()

        self.loading_table = False
        self.exam_read_only = False
        self._dashboard_subject_name = None

        layout = QVBoxLayout(self)
        layout.setSpacing(14)

        title = QLabel("RESULTS ENTRY V4.1")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title)

        # =====================================
        # FILTERS
        # =====================================
        filters_group = QGroupBox("Selection Filters")
        filters_layout = QHBoxLayout(filters_group)

        self.exam = QComboBox()
        self.exam.setMinimumWidth(180)
        self.exam.setPlaceholderText("Select Exam")

        self.class_box = QComboBox()
        self.class_box.setMinimumWidth(150)
        self.class_box.setPlaceholderText("Select Class")

        self.stream_box = QComboBox()
        self.stream_box.setMinimumWidth(140)

        self.subject = QComboBox()
        self.subject.setMinimumWidth(220)
        self.subject.setPlaceholderText("Select Subject")

        filters_layout.addWidget(QLabel("Exam"))
        filters_layout.addWidget(self.exam)
        filters_layout.addWidget(QLabel("Class"))
        filters_layout.addWidget(self.class_box)
        filters_layout.addWidget(QLabel("Stream"))
        filters_layout.addWidget(self.stream_box)
        filters_layout.addWidget(QLabel("Subject"))
        filters_layout.addWidget(self.subject)

        layout.addWidget(filters_group)

        # =====================================
        # TOP SUMMARY (V4.1)
        # =====================================
        summary_group = QGroupBox("Progress Summary")
        summary_layout = QGridLayout(summary_group)

        self.expected_value = QLabel("0")
        self.entered_value = QLabel("0")
        self.missing_value = QLabel("0")
        self.completion_value = QLabel("0.00%")

        summary_labels = (
            self.expected_value,
            self.entered_value,
            self.missing_value,
            self.completion_value,
        )

        for label in summary_labels:
            label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        summary_layout.addWidget(QLabel("Expected Students"), 0, 0)
        summary_layout.addWidget(QLabel("Entered Marks"), 0, 1)
        summary_layout.addWidget(QLabel("Missing Marks"), 0, 2)
        summary_layout.addWidget(QLabel("Completion %"), 0, 3)
        summary_layout.addWidget(self.expected_value, 1, 0)
        summary_layout.addWidget(self.entered_value, 1, 1)
        summary_layout.addWidget(self.missing_value, 1, 2)
        summary_layout.addWidget(self.completion_value, 1, 3)

        layout.addWidget(summary_group)

        # =====================================
        # MARKS TABLE
        # =====================================
        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels([
            "Admission No",
            "Student Name",
            "Marks",
        ])
        self.table.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectRows
        )
        self.table.setSelectionMode(
            QAbstractItemView.SelectionMode.SingleSelection
        )
        self.table.setEditTriggers(
            QAbstractItemView.EditTrigger.AllEditTriggers
        )
        self.table.setItemDelegateForColumn(2, MarksDelegate(self.table))
        self.table.verticalHeader().setVisible(False)
        self.table.verticalHeader().setDefaultSectionSize(44)
        self.table.itemChanged.connect(self.update_summary)

        self.lock_label = QLabel("")
        self.lock_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lock_label.hide()

        header = self.table.horizontalHeader()
        header.setSectionResizeMode(
            0,
            QHeaderView.ResizeMode.ResizeToContents,
        )
        header.setSectionResizeMode(
            1,
            QHeaderView.ResizeMode.Stretch,
        )
        header.setSectionResizeMode(
            2,
            QHeaderView.ResizeMode.Interactive,
        )
        self.table.setColumnWidth(2, 160)

        layout.addWidget(self.lock_label)
        layout.addWidget(self.table, 1)

        # =====================================
        # BOTTOM BUTTONS
        # =====================================
        buttons = QHBoxLayout()
        buttons.addStretch()

        self.save_btn = QPushButton("SAVE ALL RESULTS")
        self.save_btn.setFixedHeight(40)
        self.save_btn.clicked.connect(self.save_all)

        self.import_btn = QPushButton("IMPORT EXCEL")
        self.import_btn.clicked.connect(self.import_excel)

        self.template_btn = QPushButton("DOWNLOAD TEMPLATE")
        self.template_btn.clicked.connect(self.download_template)

        buttons.addWidget(self.import_btn)
        buttons.addWidget(self.template_btn)
        buttons.addWidget(self.save_btn)
        layout.addLayout(buttons)

        # Connect events
        self.exam.currentIndexChanged.connect(lambda _: self._on_exam_or_class_changed())
        self.class_box.currentIndexChanged.connect(lambda _: self._on_exam_or_class_changed())
        self.stream_box.currentIndexChanged.connect(lambda _: self.load_subjects())
        self.subject.currentIndexChanged.connect(lambda _: self.load_students())

        EventBus.subscribe("LEVEL_CHANGED", self.on_level_changed)
        EventBus.subscribe("STUDENTS_UPDATED", self.refresh_all)
        EventBus.subscribe("EXAMS_UPDATED", self.refresh_all)

        self.refresh_all()

    def on_level_changed(self):
        self.refresh_all()

    def load(self):
        self.refresh_all()

    def open_from_dashboard(
        self,
        exam_id,
        class_name,
        subject_name,
    ):
        class_name = str(class_name).strip()
        subject_name = str(subject_name).strip()
        self._dashboard_subject_name = subject_name
        self.load_exams()
        self.load_classes()

        exam_index = self.exam.findData(exam_id)
        class_index = self.class_box.findText(class_name)

        if exam_index < 0 or class_index < 0:
            self._clear_table()
            return

        self.exam.blockSignals(True)
        self.class_box.blockSignals(True)
        try:
            self.exam.setCurrentIndex(exam_index)
            self.class_box.setCurrentIndex(class_index)
        finally:
            self.exam.blockSignals(False)
            self.class_box.blockSignals(False)

        self.load_streams()
        self.load_subjects(load_table=False)
        selected_index = -1
        for i in range(self.subject.count()):
            item_data = self.subject.itemData(i)
            item_text = self.subject.itemText(i)
            if _subject_name_matches(item_data, subject_name) or _subject_name_matches(item_text, subject_name):
                selected_index = i
                break

        self.subject.blockSignals(True)
        try:
            if selected_index >= 0:
                self.subject.setCurrentIndex(selected_index)
            elif self.subject.count() == 0:
                self._clear_table()
        finally:
            self.subject.blockSignals(False)

        QTimer.singleShot(0, lambda: self.load_students(subject_name))

    def refresh_all(self):
        self.load_exams()
        self.load_classes()
        self.load_streams()
        self.load_subjects()

    def load_exams(self):
        combo_loaders.load_results_exams(self.exam)

    def load_classes(self):
        combo_loaders.load_classes(self.class_box)

    def load_streams(self):
        combo_loaders.load_streams(
            self.stream_box,
            class_name=self.class_box.currentText().strip(),
            exam_id=self.exam.currentData(),
            level=SystemState.get_level(),
        )

    def _on_exam_or_class_changed(self):
        self.load_streams()
        self.load_subjects()

    def load_subjects(self, load_table=True):
        load_table = bool(load_table)
        exam_id = self.exam.currentData()
        class_name = self.class_box.currentText().strip()
        stream = self.stream_box.currentData()
        level = SystemState.get_level()

        if not class_name or exam_id is None:
            self.subject.clear()
            self._clear_table()
            return

        current_subject_full = self.subject.currentText()
        current_subject_base = re.sub(r"\s*\(\d+%\)\s*$", "", current_subject_full)

        context = get_exam_context(exam_id)
        if not context:
            self.subject.clear()
            self._clear_table()
            return

        year_id, term_id = context

        rows = fetch_all("""
            SELECT 
                e.subject_name,
                COUNT(DISTINCT e.admission_no) as expected,
                (SELECT COUNT(*) FROM results r 
                 JOIN students s2 ON s2.admission_no = r.admission_no
                 WHERE UPPER(TRIM(r.subject_name)) = UPPER(TRIM(e.subject_name))
                   AND r.exam_id = ?
                   AND UPPER(TRIM(COALESCE(r.class_name, s2.class))) = UPPER(TRIM(?))
                   AND s2.level = ?
                   AND (? IS NULL OR UPPER(TRIM(COALESCE(r.stream, s2.stream, ''))) = UPPER(TRIM(?)))
                 ) as entered
            FROM enrollments e
            JOIN students es ON es.admission_no = e.admission_no
            WHERE UPPER(TRIM(e.class_name)) = UPPER(TRIM(?))
              AND e.academic_year_id=? AND e.term_id=?
              AND (? IS NULL OR UPPER(TRIM(COALESCE(es.stream, ''))) = UPPER(TRIM(?)))
            GROUP BY e.subject_name
            ORDER BY e.subject_name
        """, (exam_id, class_name, level, stream, stream,
               class_name, year_id, term_id, stream, stream))

        self.subject.blockSignals(True)
        self.subject.clear()

        for name, expected, entered in rows:
            # Hide subject only if all students already have marks (100% complete)
            if entered >= expected:
                continue
            perc = (entered / expected * 100) if expected > 0 else 0
            display_name = f"{name} ({perc:.0f}%)"
            self.subject.addItem(display_name, name)

        # Restore selection
        for i in range(self.subject.count()):
            if _subject_name_matches(self.subject.itemData(i), current_subject_base) or _subject_name_matches(self.subject.itemText(i), current_subject_base):
                self.subject.setCurrentIndex(i)
                break
        else:
            if self.subject.count() > 0:
                self.subject.setCurrentIndex(0)
            else:
                self._clear_table()

        self.subject.blockSignals(False)
        if load_table:
            self.load_students()

    def load_students(self, subject_name=None):
        exam_id = self.exam.currentData()
        class_name = self.class_box.currentText().strip()
        stream = self.stream_box.currentData()
        if isinstance(subject_name, int):
            subject_name = None

        subject_name = (
            subject_name
            or self._dashboard_subject_name
            or self.subject.currentData()
            or re.sub(r"\s*\(\d+%\)\s*$", "", self.subject.currentText()).strip()
        )
        level = SystemState.get_level()

        if exam_id is None or not class_name or not subject_name:
            self._clear_table()
            return

        self._update_exam_lock_state()

        context = get_exam_context(exam_id)
        if not context:
            self._clear_table()
            return

        year_id, term_id = context

        student_rows = fetch_all("""
            SELECT DISTINCT
                s.admission_no,
                s.full_name,
                TRIM(COALESCE(s.stream, ''))
            FROM enrollments e
            JOIN students s ON s.admission_no = e.admission_no
            WHERE UPPER(TRIM(e.subject_name)) = UPPER(TRIM(?))
              AND s.level = ?
              AND COALESCE(s.status, 'ACTIVE') = 'ACTIVE'
              AND COALESCE(e.is_active, 1) = 1
              AND UPPER(TRIM(e.class_name)) = UPPER(TRIM(?))
              AND e.academic_year_id = ?
              AND e.term_id = ?
              AND (? IS NULL OR UPPER(TRIM(COALESCE(s.stream, ''))) = UPPER(TRIM(?)))
            ORDER BY s.full_name
        """, (subject_name, level, class_name, year_id, term_id, stream, stream))

        result_rows = fetch_all("""
            SELECT
                admission_no,
                marks
            FROM results
            WHERE exam_id = ?
              AND UPPER(TRIM(subject_name)) = UPPER(TRIM(?))
              AND UPPER(TRIM(COALESCE(class_name, ''))) = UPPER(TRIM(?))
              AND (? IS NULL OR UPPER(TRIM(COALESCE(stream, ''))) = UPPER(TRIM(?)))
        """, (exam_id, subject_name, class_name, stream, stream))

        marks_by_admission = {
            admission_no: marks
            for admission_no, marks in result_rows
        }
        rows = [
            (admission_no, full_name, student_stream, marks_by_admission.get(admission_no))
            for admission_no, full_name, student_stream in student_rows
        ]

        self.loading_table = True
        self.table.setRowCount(len(rows))

        read_only_flags = (
            Qt.ItemFlag.ItemIsEnabled
            | Qt.ItemFlag.ItemIsSelectable
        )
        marks_flags = read_only_flags
        if not self.exam_read_only:
            marks_flags |= Qt.ItemFlag.ItemIsEditable

        for row_index, (admission_no, full_name, student_stream, marks) in enumerate(rows):
            admission_item = QTableWidgetItem(admission_no)
            admission_item.setFlags(read_only_flags)
            admission_item.setData(Qt.ItemDataRole.UserRole, student_stream)

            name_item = QTableWidgetItem(full_name or "")
            name_item.setFlags(read_only_flags)

            marks_item = QTableWidgetItem(
                "" if marks is None else str(marks)
            )
            marks_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            marks_item.setFlags(marks_flags)

            self.table.setItem(row_index, 0, admission_item)
            self.table.setItem(row_index, 1, name_item)
            self.table.setItem(row_index, 2, marks_item)
            if self.exam_read_only:
                self.table.closePersistentEditor(marks_item)
            else:
                self.table.openPersistentEditor(marks_item)

        self.loading_table = False
        self.update_summary()
        self._dashboard_subject_name = None

    def save_all(self):
        exam_id = self.exam.currentData()
        subject_name = self.subject.currentData()

        if exam_id is None or not subject_name:
            show_error(self, "Select an exam, class and subject.", title="Missing Filters")
            return

        if self._is_selected_exam_completed():
            show_error(
                self,
                "This exam is completed and read-only. Archived exams can still be viewed and exported.",
                title="Completed Exam"
            )
            return

        marks_to_save = []
        invalid_rows = []
        class_name = self.class_box.currentText().strip()

        for row in range(self.table.rowCount()):
            admission_item = self.table.item(row, 0)
            marks_item = self.table.item(row, 2)

            if admission_item is None: continue

            marks_text = marks_item.text().strip() if marks_item is not None else ""
            if not marks_text: continue

            try:
                marks = int(marks_text)
                if not (0 <= marks <= 100): raise ValueError()
            except ValueError:
                invalid_rows.append(row + 1)
                continue

            marks_to_save.append((
                admission_item.text(), marks,
                admission_item.data(Qt.ItemDataRole.UserRole) or "",
            ))

        if invalid_rows:
            show_error(self, f"Check row(s): {', '.join(map(str, invalid_rows))}", title="Invalid Marks")
            return

        try:
            with get_cursor(commit=True) as cur:
                for admission_no, marks, stream in marks_to_save:
                    cur.execute("""
                        INSERT INTO results (admission_no, subject_name, marks, exam_id, class_name, stream)
                        VALUES (?, ?, ?, ?, ?, ?)
                        ON CONFLICT(admission_no, subject_name, exam_id)
                        DO UPDATE SET marks = excluded.marks,
                                      class_name = excluded.class_name,
                                      stream = excluded.stream
                    """, (admission_no, subject_name, marks, exam_id, class_name, stream))
        except Exception as e:
            QMessageBox.critical(self, "Database Error", f"An unexpected error occurred while saving results: {e}")
            return

        # V4.1: Automatically refresh subject completion
        self.load_subjects()
        EventBus.emit("RESULTS_UPDATED")

        show_info(self, "Results Saved Successfully")

    def update_summary(self, _item=None):
        if self.loading_table:
            return

        expected = self.table.rowCount()
        entered = 0

        for row in range(expected):
            item = self.table.item(row, 2)
            text = item.text().strip() if item is not None else ""
            if text: entered += 1

        missing = expected - entered
        completion = (entered / expected * 100) if expected else 0

        self.expected_value.setText(str(expected))
        self.entered_value.setText(str(entered))
        self.missing_value.setText(str(missing))
        self.completion_value.setText(f"{completion:.2f}%")

    # =========================
    # EXCEL FRAMEWORK
    # =========================

    def download_template(self):
        """Generate a branded Excel template with pre-filled student names and admission numbers."""
        exam_id = self.exam.currentData()
        subject_name = self.subject.currentData()
        class_name = self.class_box.currentText().strip()

        if exam_id is None or not subject_name or not class_name:
            show_error(self, "Please select Exam, Class, and Subject first.", title="Missing Filters")
            return

        level = SystemState.get_level()
        context = get_exam_context(exam_id)
        if not context:
            show_error(self, "Invalid exam context.", title="Error")
            return
        year_id, term_id = context

        # Fetch enrolled students for this subject/class/term
        students = fetch_all("""
            SELECT DISTINCT s.admission_no, s.full_name
            FROM enrollments e
            JOIN students s ON s.admission_no = e.admission_no
            WHERE UPPER(TRIM(e.subject_name)) = UPPER(TRIM(?))
              AND s.level = ?
              AND COALESCE(s.status, 'ACTIVE') = 'ACTIVE'
              AND COALESCE(e.is_active, 1) = 1
              AND UPPER(TRIM(e.class_name)) = UPPER(TRIM(?))
              AND e.academic_year_id = ?
              AND e.term_id = ?
            ORDER BY s.full_name
        """, (subject_name, level, class_name, year_id, term_id))

        if not students:
            show_error(self, "No students enrolled for this subject and class.", title="No Data")
            return

        # Exam details (school profile is now fetched inside write_branding_header)
        exam_info = fetch_one("""
            SELECT e.exam_name, t.term_name, y.year_name
            FROM exams e
            JOIN terms t ON t.id = e.term_id
            JOIN academic_years y ON y.id = t.academic_year_id
            WHERE e.id = ?
        """, (exam_id,))
        if exam_info:
            exam_name, term_name, year_name = exam_info
        else:
            exam_name = "SELECTED EXAM"
            term_name = "SELECTED TERM"
            year_name = "SELECTED YEAR"

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Marks Entry"

        # ===== Styling helpers =====
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ===== Shared school branding (rows 1-3: name, contact, title + logo) =====
        # Uses the same helper as download_template()/export_to_excel() in
        # excel_utils.py, so every generated workbook -- templates, exports,
        # and this marks-entry sheet -- shares one school-header source of
        # truth (name, contact line, brand-blue title, and logo).
        blue_fill, white_font, _ = excel_utils.write_branding_header(
            ws, 3, f"Examination Marks Entry Form \u2013 {exam_name}"
        )
        row = 4

        # ===== Subtitle (class/subject) =====
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = f"Class: {class_name}  |  Subject: {subject_name}  |  Level: {level}  |  Term: {term_name}  |  Year: {year_name}"
        cell.font = Font(bold=True, size=12)
        cell.alignment = center_align
        row += 1

        row += 1  # blank line

        # ===== Instructions =====
        instructions = [
            "INSTRUCTIONS:",
            "1. Do not change the column headers.",
            "2. Enter marks between 0 and 100.",
            "3. Admission numbers are pre-filled – do not modify.",
            "4. Start data entry from the row after the header.",
        ]
        for line in instructions:
            ws.merge_cells(f'A{row}:C{row}')
            cell = ws[f'A{row}']
            cell.value = line
            cell.alignment = left_align
            cell.font = Font(size=10)
            row += 1

        row += 1  # blank line

        # ===== Header row =====
        # Same navy-fill / white-text convention as excel_utils' templates
        # and exports, plus this page's own thin-border touch.
        headers = ["Admission No*", "Student Name", "Marks (0-100)*"]
        header_row = row
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.fill = blue_fill
            cell.font = white_font
            cell.alignment = center_align
            cell.border = thin_border

        # ===== Student data =====
        data_row = header_row + 1
        for admission_no, full_name in students:
            ws.cell(row=data_row, column=1, value=admission_no)
            ws.cell(row=data_row, column=2, value=full_name)
            ws.cell(row=data_row, column=3, value="")  # empty marks
            data_row += 1

        # ===== Apply borders to data rows =====
        for r in range(header_row + 1, data_row):
            for c in range(1, 4):
                cell = ws.cell(row=r, column=c)
                cell.border = thin_border
                cell.alignment = center_align if c != 2 else left_align

        # ===== Column widths =====
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 18

        # ===== Freeze header row =====
        ws.freeze_panes = f'A{header_row + 1}'

        # ===== Save file =====
        default_name = f"marks_template_{class_name}_{subject_name}.xlsx"
        save_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Template",
            default_name,
            "Excel Files (*.xlsx)"
        )
        if not save_path:
            return
        if not save_path.lower().endswith(".xlsx"):
            save_path += ".xlsx"

        try:
            wb.save(save_path)
            show_info(self, f"Template saved to {save_path}")
        except Exception as e:
            show_error(self, f"Failed to save template: {e}", title="Error")

    def import_excel(self):
        exam_id = self.exam.currentData()
        subject_name = self.subject.currentData()

        if not (exam_id and subject_name):
            show_error(self, "Select Exam and Subject first")
            return

        if self._is_selected_exam_completed():
            show_error(
                self,
                "This exam is completed and read-only. Archived exams cannot accept new imports.",
                title="Completed Exam"
            )
            return

        class_name = self.class_box.currentText().strip()
        if not class_name:
            show_error(self, "Please select a class.")
            return

        path = excel_utils.get_import_file(self)
        if not path:
            return

        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            sheet = wb.active

            # Find header row (contains "Admission No*" etc.)
            header_row = None
            for row_num in range(1, 20):
                cell_val = sheet.cell(row=row_num, column=1).value
                if cell_val and "Admission" in str(cell_val):
                    header_row = row_num
                    break
            if header_row is None:
                raise ValueError(
                    "Could not detect the template's header row. "
                    "Please use a template downloaded from this system."
                )

            data_start_row = header_row + 1
            rows = list(sheet.iter_rows(min_row=data_start_row, values_only=True))

            context = get_exam_context(exam_id)
            if not context:
                raise ValueError("Selected exam does not exist.")
            year_id, term_id = context

            imported = 0
            with get_cursor(commit=True) as cur:
                for row in rows:
                    if not row or not row[0]:
                        continue
                    adm_no = str(row[0]).strip()
                    # Marks are in column 3 (index 2)
                    marks_raw = row[2] if len(row) > 2 else None

                    if marks_raw is None:
                        continue

                    try:
                        marks_int = int(marks_raw)
                        if not (0 <= marks_int <= 100):
                            raise ValueError("Marks must be between 0 and 100")
                    except (ValueError, TypeError) as e:
                        print(f"[WARNING] Skipping invalid marks for {adm_no}: {marks_raw} ({e})")
                        continue

                    cur.execute("""
                        SELECT 1 FROM enrollments
                        WHERE admission_no=? AND subject_name=? AND class_name=? AND academic_year_id=? AND term_id=?
                    """, (adm_no, subject_name, class_name, year_id, term_id))

                    if cur.fetchone():
                        try:
                            cur.execute(
                                "SELECT TRIM(COALESCE(stream, '')) FROM students WHERE admission_no=?",
                                (adm_no,),
                            )
                            stream_row = cur.fetchone()
                            student_stream = stream_row[0] if stream_row else ""
                            cur.execute("""
                                INSERT INTO results (admission_no, subject_name, marks, exam_id, class_name, stream)
                                VALUES (?, ?, ?, ?, ?, ?)
                                ON CONFLICT(admission_no, subject_name, exam_id) DO UPDATE SET
                                    marks=excluded.marks,
                                    class_name=excluded.class_name,
                                    stream=excluded.stream
                            """, (adm_no, subject_name, marks_int, exam_id, class_name, student_stream))
                            imported += 1
                        except Exception as e:
                            print(f"[ERROR] Failed to import result for '{adm_no}': {e}")
                            continue

            self.load_students()
            self.load_subjects()
            EventBus.emit("RESULTS_UPDATED")
            QMessageBox.information(self, "Import Complete", f"Imported {imported} marks.")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Import failed: {e}")

    def _is_selected_exam_completed(self):
        exam_id = self.exam.currentData()
        if exam_id is None:
            return False

        row = fetch_one(
            "SELECT status FROM exams WHERE id=?",
            (exam_id,),
        )
        return bool(row and row[0] == "COMPLETED")

    def _update_exam_lock_state(self):
        self.exam_read_only = self._is_selected_exam_completed()
        self.save_btn.setEnabled(not self.exam_read_only)
        self.import_btn.setEnabled(not self.exam_read_only)

        if self.exam_read_only:
            self.lock_label.setText(
                "COMPLETED EXAM - archived results are read-only."
            )
            self.lock_label.show()
            self.table.setEditTriggers(
                QAbstractItemView.EditTrigger.NoEditTriggers
            )
        else:
            self.lock_label.hide()
            self.table.setEditTriggers(
                QAbstractItemView.EditTrigger.AllEditTriggers
            )

    def _clear_table(self):
        self._update_exam_lock_state()
        self.loading_table = True
        self.table.setRowCount(0)
        self.loading_table = False
        self.update_summary()
