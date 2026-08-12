import logging

from PySide6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QPushButton,
    QComboBox,
    QTableWidget,
    QTableWidgetItem,
    QMessageBox,
    QAbstractItemView,
    QHeaderView,
    QLabel,
    QFrame,
    QScrollArea,
    QSizePolicy,
    QCheckBox,
    QDialog,
    QDialogButtonBox,
)

from PySide6.QtCore import Qt

from db_utils import get_cursor, fetch_all
from system_state import SystemState
from event_bus import EventBus
from class_utils import get_classes
from ui_helpers import show_error, show_info, get_subject_short_name
import combo_loaders

logger = logging.getLogger(__name__)


class EnrollmentPage(QWidget):

    def __init__(self):
        super().__init__()

        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setFrameShape(QFrame.NoFrame)
        root.addWidget(self.scroll_area)

        self.content_widget = QWidget()
        self.scroll_area.setWidget(self.content_widget)

        self.content_layout = QVBoxLayout(self.content_widget)
        self.content_layout.setContentsMargins(20, 20, 20, 20)
        self.content_layout.setSpacing(12)

        # =========================
        # FILTERS
        # =========================

        filters_layout = QHBoxLayout()

        # Year
        self.year_box = QComboBox()
        self.year_box.currentIndexChanged.connect(self.load_terms)

        # Term
        self.term_box = QComboBox()
        self.term_box.currentIndexChanged.connect(self.on_filter_changed)

        # Class
        self.class_box = QComboBox()
        self.class_box.addItems(["-- Select Class --"] + get_classes())
        self.class_box.currentIndexChanged.connect(self.load_students)

        filters_layout.addWidget(QLabel("Year:"))
        filters_layout.addWidget(self.year_box)
        filters_layout.addWidget(QLabel("Term:"))
        filters_layout.addWidget(self.term_box)
        filters_layout.addWidget(QLabel("Class:"))
        filters_layout.addWidget(self.class_box)
        
        self.import_btn = QPushButton("Import Excel")
        self.import_btn.setObjectName("workflowSecondary")
        self.import_btn.clicked.connect(self.import_excel)

        self.export_btn = QPushButton("Export Excel")
        self.export_btn.setObjectName("workflowSecondary")
        self.export_btn.clicked.connect(self.export_excel)

        self.template_btn = QPushButton("Download Template")
        self.template_btn.setObjectName("workflowSecondary")
        self.template_btn.clicked.connect(self.download_template)

        filters_layout.addWidget(self.template_btn)
        filters_layout.addWidget(self.import_btn)
        filters_layout.addWidget(self.export_btn)
        
        filters_layout.addStretch()

        self.content_layout.addLayout(filters_layout)

        # =========================
        # MODE / PREVIEW
        # =========================

        mode_layout = QHBoxLayout()
        self.enrollment_mode_checkbox = QCheckBox("Enrollment Mode")
        self.enrollment_mode_checkbox.setChecked(True)
        self.enrollment_mode_checkbox.toggled.connect(self.set_enrollment_mode)
        mode_layout.addWidget(self.enrollment_mode_checkbox)

        self.preview_label = QLabel("Preview mode: changes are disabled until Enrollment Mode is enabled.")
        self.preview_label.setWordWrap(True)
        self.preview_label.setProperty("variant", "muted")
        mode_layout.addWidget(self.preview_label, 1)
        mode_layout.addStretch()
        self.content_layout.addLayout(mode_layout)

        # =========================
        # TABLE AREA
        # =========================

        self.enrollment_table = QTableWidget()
        self.enrollment_table.setColumnCount(1)
        self.enrollment_table.setHorizontalHeaderLabels(["Student Name"])
        self.enrollment_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.enrollment_table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.enrollment_table.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.enrollment_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.enrollment_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.enrollment_table.setMinimumHeight(400)
        self.content_layout.addWidget(QLabel("DYNAMIC ENROLLMENT GRID"))

        # Per-subject bulk-enroll/unenroll buttons, rebuilt each time the
        # subject list changes (see _rebuild_subject_bulk_rows).
        self.subject_bulk_row = QHBoxLayout()
        self.subject_bulk_row.addWidget(QLabel("Enroll all in:"))
        self.subject_bulk_row.addStretch()
        self.content_layout.addLayout(self.subject_bulk_row)

        self.subject_unenroll_row = QHBoxLayout()
        self.subject_unenroll_row.addWidget(QLabel("Unenroll all in:"))
        self.subject_unenroll_row.addStretch()
        self.content_layout.addLayout(self.subject_unenroll_row)

        self.content_layout.addWidget(self.enrollment_table)

        # =========================
        # SAVE BUTTON
        # =========================

        action_row = QHBoxLayout()
        action_row.addStretch()

        self.copy_previous_btn = QPushButton("Copy From Previous Term")
        self.copy_previous_btn.setObjectName("workflowSecondary")
        self.copy_previous_btn.setFixedHeight(40)
        self.copy_previous_btn.clicked.connect(self.copy_from_previous_term)

        self.enroll_all_btn = QPushButton("Enroll All")
        self.enroll_all_btn.setObjectName("workflowWarning")
        self.enroll_all_btn.setFixedHeight(40)
        self.enroll_all_btn.clicked.connect(self.enroll_all)

        self.unenroll_all_btn = QPushButton("Unenroll All")
        self.unenroll_all_btn.setObjectName("workflowWarning")
        self.unenroll_all_btn.setFixedHeight(40)
        self.unenroll_all_btn.clicked.connect(self.unenroll_all)

        self.save_btn = QPushButton("Save Enrollments")
        self.save_btn.setObjectName("workflowPrimary")
        self.save_btn.setFixedHeight(40)
        self.save_btn.clicked.connect(self.save_enrollments)

        action_row.addWidget(self.copy_previous_btn)
        action_row.addWidget(self.enroll_all_btn)
        action_row.addWidget(self.unenroll_all_btn)
        action_row.addWidget(self.save_btn)
        self.content_layout.addLayout(action_row)

        # =========================
        # INITIAL LOAD
        # =========================

        EventBus.subscribe("LEVEL_CHANGED", self.on_level_changed)

        self.student_list = []
        self.subject_list = []

        self.load_years()
        self.set_enrollment_mode(True)

    # =========================
    # EVENT HANDLERS
    # =========================

    def on_level_changed(self):
        # Reload classes for the new level. If the combo's selected *index*
        # happens to land on the same position as before, currentIndexChanged
        # won't fire even though the underlying class/level has changed, so
        # self.student_list/self.subject_list would otherwise go stale while
        # the grid looks empty. Reset state explicitly and always reload.
        self.student_list = []
        self.subject_list = []
        combo_loaders.load_classes(self.class_box, placeholder="-- Select Class --")
        self.clear_tables()
        self.load_students()

    def on_filter_changed(self):
        self.load_enrollment_data()

    def set_enrollment_mode(self, enabled):
        self.enrollment_mode_checkbox.setChecked(enabled)
        self.preview_label.setText(
            "Preview mode: changes are disabled until Enrollment Mode is enabled."
            if not enabled
            else "Enrollment mode is active. You can edit the student subject list."
        )

        self.save_btn.setEnabled(enabled)
        self.import_btn.setEnabled(enabled)
        self.enroll_all_btn.setEnabled(enabled)
        self.unenroll_all_btn.setEnabled(enabled)
        self.copy_previous_btn.setEnabled(enabled)
        self.enrollment_table.setEnabled(enabled)
        for row in (self.subject_bulk_row, self.subject_unenroll_row):
            for i in range(row.count()):
                widget = row.itemAt(i).widget()
                if isinstance(widget, QPushButton):
                    widget.setEnabled(enabled)

    # =========================
    # DATA LOADING
    # =========================

    def load_years(self):
        combo_loaders.load_years(self.year_box)
        self.load_terms()

    def load_terms(self):
        # Repopulating term_box would otherwise fire currentIndexChanged
        # (-> on_filter_changed) mid-update, causing load_enrollment_data to
        # run once with a stale/old term before settling on the new one.
        # Block signals here and trigger a single, consistent reload after.
        self.term_box.blockSignals(True)
        combo_loaders.load_terms(self.term_box, self.year_box.currentData())
        self.term_box.blockSignals(False)
        self.load_enrollment_data()

    def load_students(self):
        class_name = self.class_box.currentText()
        self.student_list = []
        self.subject_list = []

        if class_name and class_name != "-- Select Class --":
            self.student_list = [
                (row[0], row[1])
                for row in fetch_all(
                    """
                    SELECT admission_no, full_name
                    FROM students
                    WHERE class=? AND level=?
                    ORDER BY full_name
                    """,
                    (class_name, SystemState.get_level()),
                )
            ]

        self.load_enrollment_data()

    def load_enrollment_data(self):
        self.enrollment_table.clear()

        year_id = self.year_box.currentData()
        term_id = self.term_box.currentData()
        class_name = self.class_box.currentText()

        if not (year_id and term_id and class_name and class_name != "-- Select Class --"):
            self._reset_table_to_empty()
            self._rebuild_subject_bulk_rows()
            return

        self.subject_list = [
            row[0]
            for row in fetch_all(
                """
                SELECT subject_name
                FROM subjects
                WHERE level=?
                ORDER BY subject_name
                """,
                (SystemState.get_level(),),
            )
        ]

        enrollments = {
            (row[0], row[1])
            for row in fetch_all(
                """
                SELECT e.admission_no, e.subject_name
                FROM enrollments e
                JOIN students s ON s.admission_no = e.admission_no
                WHERE e.academic_year_id=? AND e.term_id=?
                  AND e.class_name=?
                  AND s.level=?
                """,
                (year_id, term_id, class_name, SystemState.get_level()),
            )
        }

        self.enrollment_table.setRowCount(len(self.student_list))
        self.enrollment_table.setColumnCount(len(self.subject_list) + 1)

        headers = ["Student"] + [get_subject_short_name(subject) for subject in self.subject_list]
        self.enrollment_table.setHorizontalHeaderLabels(headers)
        self.enrollment_table.verticalHeader().setVisible(False)

        for row_index, (admission_no, full_name) in enumerate(self.student_list):
            student_item = QTableWidgetItem(full_name)
            student_item.setFlags(student_item.flags() & ~Qt.ItemIsEditable)
            student_item.setData(Qt.UserRole, admission_no)
            student_item.setToolTip(admission_no)
            self.enrollment_table.setItem(row_index, 0, student_item)

            for col_index, subject_name in enumerate(self.subject_list, start=1):
                checkbox = QTableWidgetItem()
                checkbox.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
                checkbox.setCheckState(
                    Qt.Checked if (admission_no, subject_name) in enrollments else Qt.Unchecked
                )
                checkbox.setData(Qt.UserRole, subject_name)
                self.enrollment_table.setItem(row_index, col_index, checkbox)

        self.enrollment_table.resizeColumnsToContents()
        self.enrollment_table.resizeRowsToContents()
        self._rebuild_subject_bulk_rows()

    def _rebuild_subject_bulk_rows(self):
        """Rebuild the rows of per-subject 'enroll everyone' / 'unenroll
        everyone' buttons to match the current subject_list. Kept separate
        from enroll_all/unenroll_all so a single subject can be bulk-toggled
        without touching every other column."""
        self._rebuild_subject_button_row(self.subject_bulk_row, self.enroll_all_for_subject)
        self._rebuild_subject_button_row(self.subject_unenroll_row, self.unenroll_all_for_subject)

    def _rebuild_subject_button_row(self, row_layout, handler):
        """Clear and repopulate a per-subject button row (skip the leading
        label and trailing stretch, which stay fixed) using the given
        click handler."""
        while row_layout.count() > 2:
            item = row_layout.takeAt(1)
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()

        for subject_name in self.subject_list:
            btn = QPushButton(get_subject_short_name(subject_name))
            btn.setObjectName("workflowSecondary")
            btn.setEnabled(self.enrollment_mode_checkbox.isChecked())
            btn.clicked.connect(
                lambda checked=False, s=subject_name: handler(s)
            )
            row_layout.insertWidget(row_layout.count() - 1, btn)

    def clear_tables(self):
        self._reset_table_to_empty()

    def _reset_table_to_empty(self):
        self.enrollment_table.setRowCount(0)
        self.enrollment_table.setColumnCount(1)
        self.enrollment_table.setHorizontalHeaderLabels(["Student Name"])

    # =========================
    # SAVE
    # =========================

    def save_enrollments(self):
        year_id = self.year_box.currentData()
        term_id = self.term_box.currentData()
        class_name = self.class_box.currentText()

        if not (year_id and term_id and class_name and class_name != "-- Select Class --"):
            show_error(self, "Please select Year, Term, and Class before saving enrollments.")
            return

        enrolled_subjects = []
        for row_index, (admission_no, _) in enumerate(self.student_list):
            for col_index, subject_name in enumerate(self.subject_list, start=1):
                item = self.enrollment_table.item(row_index, col_index)
                if item is None:
                    continue
                if item.checkState() == Qt.Checked:
                    enrolled_subjects.append((admission_no, subject_name))

        admission_numbers = [student_id for student_id, _ in self.student_list]

        try:
            with get_cursor(commit=True) as cur:
                if admission_numbers:
                    placeholders = ",".join("?" for _ in admission_numbers)
                    cur.execute(f"""
                        DELETE FROM enrollments
                        WHERE academic_year_id=?
                          AND term_id=?
                          AND class_name=?
                          AND admission_no IN ({placeholders})
                    """, (year_id, term_id, class_name, *admission_numbers))

                if enrolled_subjects:
                    cur.executemany(
                        """
                        INSERT OR REPLACE INTO enrollments(admission_no, subject_name, class_name, academic_year_id, term_id)
                        VALUES (?, ?, ?, ?, ?)
                        """,
                        [
                            (admission_no, subject_name, class_name, year_id, term_id)
                            for admission_no, subject_name in enrolled_subjects
                        ],
                    )

            show_info(self, "Enrollments saved successfully.")

        except Exception:
            logger.exception("Failed to save enrollments for class=%s year=%s term=%s", class_name, year_id, term_id)
            QMessageBox.critical(self, "Error", "An unexpected error occurred while saving enrollments.")

        self.load_enrollment_data()

    def enroll_all(self):
        year_id = self.year_box.currentData()
        term_id = self.term_box.currentData()
        class_name = self.class_box.currentText()

        if not (year_id and term_id and class_name and class_name != "-- Select Class --"):
            show_error(self, "Please select Year, Term, and Class before using Enroll All.")
            return

        if not self.student_list or not self.subject_list:
            show_error(self, "No students or subjects found for the selected class.")
            return

        reply = QMessageBox.question(
            self,
            "Confirm Enroll All",
            f"This will enroll all {len(self.student_list)} students in {class_name} "
            f"into all {len(self.subject_list)} subjects for {self.year_box.currentText()} "
            f"{self.term_box.currentText()}, replacing the current enrollment selection "
            "for this class and term. This cannot be undone.\n\nContinue?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return

        for row_index in range(self.enrollment_table.rowCount()):
            for col_index in range(1, self.enrollment_table.columnCount()):
                item = self.enrollment_table.item(row_index, col_index)
                if item is not None:
                    item.setCheckState(Qt.Checked)

        self.save_enrollments()

    def unenroll_all(self):
        year_id = self.year_box.currentData()
        term_id = self.term_box.currentData()
        class_name = self.class_box.currentText()

        if not (year_id and term_id and class_name and class_name != "-- Select Class --"):
            show_error(self, "Please select Year, Term, and Class before using Unenroll All.")
            return

        if not self.student_list or not self.subject_list:
            show_error(self, "No students or subjects found for the selected class.")
            return

        reply = QMessageBox.question(
            self,
            "Confirm Unenroll All",
            f"This will unenroll all {len(self.student_list)} students in {class_name} "
            f"from all {len(self.subject_list)} subjects for {self.year_box.currentText()} "
            f"{self.term_box.currentText()}, clearing the current enrollment selection "
            "for this class and term. This cannot be undone.\n\nContinue?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return

        for row_index in range(self.enrollment_table.rowCount()):
            for col_index in range(1, self.enrollment_table.columnCount()):
                item = self.enrollment_table.item(row_index, col_index)
                if item is not None:
                    item.setCheckState(Qt.Unchecked)

        self.save_enrollments()

    def unenroll_all_for_subject(self, subject_name):
        """Uncheck every student's box for a single subject only, leaving
        all other subjects untouched. Scoped alternative to unenroll_all()."""
        year_id = self.year_box.currentData()
        term_id = self.term_box.currentData()
        class_name = self.class_box.currentText()

        if not (year_id and term_id and class_name and class_name != "-- Select Class --"):
            show_error(self, "Please select Year, Term, and Class first.")
            return

        if subject_name not in self.subject_list:
            return

        col_index = self.subject_list.index(subject_name) + 1
        display_name = get_subject_short_name(subject_name)

        reply = QMessageBox.question(
            self,
            "Confirm Unenroll All",
            f"Unenroll all {len(self.student_list)} students in {class_name} "
            f"from {display_name}? This saves immediately.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return

        for row_index in range(self.enrollment_table.rowCount()):
            item = self.enrollment_table.item(row_index, col_index)
            if item is not None:
                item.setCheckState(Qt.Unchecked)

        self.save_enrollments()

    def enroll_all_for_subject(self, subject_name):
        """Check every student's box for a single subject only, leaving all
        other subjects untouched. Scoped alternative to enroll_all()."""
        year_id = self.year_box.currentData()
        term_id = self.term_box.currentData()
        class_name = self.class_box.currentText()

        if not (year_id and term_id and class_name and class_name != "-- Select Class --"):
            show_error(self, "Please select Year, Term, and Class first.")
            return

        if subject_name not in self.subject_list:
            return

        col_index = self.subject_list.index(subject_name) + 1
        display_name = get_subject_short_name(subject_name)

        reply = QMessageBox.question(
            self,
            "Confirm Enroll All",
            f"Enroll all {len(self.student_list)} students in {class_name} "
            f"into {display_name}? This saves immediately.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return

        for row_index in range(self.enrollment_table.rowCount()):
            item = self.enrollment_table.item(row_index, col_index)
            if item is not None:
                item.setCheckState(Qt.Checked)

        self.save_enrollments()

    def copy_from_previous_term(self):
        """Pre-check the current grid with enrollments from another term for
        the same class. Purely in-memory: nothing is written to the database
        until the user clicks Save Enrollments, so it's fully reviewable and
        reversible before it's committed."""
        year_id = self.year_box.currentData()
        term_id = self.term_box.currentData()
        class_name = self.class_box.currentText()

        if not (year_id and term_id and class_name and class_name != "-- Select Class --"):
            show_error(self, "Please select Year, Term, and Class first.")
            return

        # NOTE: assumes lookup tables named academic_years(id, year_name) and
        # terms(id, term_name) for display labels. Adjust the column/table
        # names below if the actual schema differs.
        candidates = fetch_all(
            """
            SELECT DISTINCT e.academic_year_id, e.term_id, ay.year_name, t.term_name
            FROM enrollments e
            JOIN academic_years ay ON ay.id = e.academic_year_id
            JOIN terms t ON t.id = e.term_id
            WHERE e.class_name = ?
              AND NOT (e.academic_year_id = ? AND e.term_id = ?)
            ORDER BY e.academic_year_id DESC, e.term_id DESC
            """,
            (class_name, year_id, term_id),
        )

        if not candidates:
            show_info(self, "No enrollment data found for this class in any other term.")
            return

        labels = [f"{year_name} - {term_name}" for _, _, year_name, term_name in candidates]

        dialog = QDialog(self)
        dialog.setWindowTitle("Copy From Previous Term")
        layout = QVBoxLayout(dialog)
        layout.addWidget(QLabel(f"Copy enrollments into {class_name} from:"))
        term_picker = QComboBox()
        term_picker.addItems(labels)
        layout.addWidget(term_picker)
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)

        if dialog.exec() != QDialog.Accepted:
            return

        source_year_id, source_term_id, source_year_name, source_term_name = candidates[term_picker.currentIndex()]

        source_enrollments = {
            (row[0], row[1])
            for row in fetch_all(
                """
                SELECT admission_no, subject_name
                FROM enrollments
                WHERE academic_year_id=? AND term_id=? AND class_name=?
                """,
                (source_year_id, source_term_id, class_name),
            )
        }

        if not source_enrollments:
            show_info(self, f"No enrollment records found for {source_year_name} - {source_term_name}.")
            return

        applied = 0
        for row_index, (admission_no, _) in enumerate(self.student_list):
            for col_index, subject_name in enumerate(self.subject_list, start=1):
                if (admission_no, subject_name) in source_enrollments:
                    item = self.enrollment_table.item(row_index, col_index)
                    if item is not None and item.checkState() != Qt.Checked:
                        item.setCheckState(Qt.Checked)
                        applied += 1

        show_info(
            self,
            f"Copied {applied} enrollment(s) from {source_year_name} - {source_term_name} "
            "into the grid. Nothing has been saved yet — review the selections and click "
            "Save Enrollments to commit them.",
        )

    # =========================
    # EXCEL FRAMEWORK
    # =========================

    def download_template(self):
        import excel_utils
        year = self.year_box.currentText().strip() or "SELECTED YEAR"
        term = self.term_box.currentText().strip() or "SELECTED TERM"
        level = SystemState.get_level()

        subjects = [
            row[0]
            for row in fetch_all(
                "SELECT subject_name FROM subjects WHERE level=? ORDER BY subject_name",
                (level,),
            )
        ]
        if not subjects:
            show_error(self, "No subjects are configured for the selected level yet.")
            return

        # Wide format: one column per subject, admission numbers listed
        # underneath (ragged — subjects have different numbers of students).
        excel_utils.download_template(
            self,
            "enrollment_template.xlsx",
            f"STUDENT SUBJECT ENROLLMENT FORM - {year} {term}",
            subjects,
            instructions=[
                f"1. Template generated for Year: {year}, Term: {term}, Level: {level}.",
                "2. Do not modify the subject headers in Row 10.",
                "3. Starting from Row 12, list the Admission No of each student under the subject(s) they take.",
                "4. Leave a cell blank if a student is not taking that subject.",
                "5. Columns do not need to have the same number of entries.",
                "6. Admission numbers must already exist in the system and belong to the selected class and level.",
            ],
            samples=None,
        )

    def export_excel(self):
        import excel_utils
        year_id = self.year_box.currentData()
        term_id = self.term_box.currentData()
        if not (year_id and term_id):
            show_error(self, "Select Year and Term first")
            return
            
        data = fetch_all("""
            SELECT admission_no, subject_name
            FROM enrollments 
            WHERE academic_year_id=? AND term_id=? AND class_name=?
        """, (year_id, term_id, self.class_box.currentText()))
        
        excel_utils.export_to_excel(
            self, 
            "enrollments.xlsx", 
            ["Admission No", "Subject Name"],
            data
        )

    def _read_long_format_pairs(self, sheet, data_start_row):
        """Legacy 2-column template: Admission No | Subject Name, one pair per row."""
        pairs = []
        for row_number, row in enumerate(sheet.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
            if not row or len(row) < 2 or row[0] in (None, "") or row[1] in (None, ""):
                continue
            pairs.append((f"Row {row_number}", row[0], row[1]))
        return pairs

    def _read_wide_format_pairs(self, sheet, header_row, data_start_row):
        """One column per subject; admission numbers listed underneath.
        Columns are read independently and are allowed to be ragged (different
        numbers of entries per subject) — a blank cell just ends that
        student's entry, not the whole column, so gaps mid-column are
        tolerated rather than treated as the end of data."""
        pairs = []
        for col_index, subject_raw in enumerate(header_row):
            if subject_raw in (None, ""):
                continue
            for row_number, row in enumerate(
                sheet.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row
            ):
                if col_index >= len(row):
                    continue
                admission_raw = row[col_index]
                if admission_raw in (None, ""):
                    continue
                pairs.append((f"Row {row_number}, column '{subject_raw}'", admission_raw, subject_raw))
        return pairs

    def import_excel(self):
        import excel_utils
        import openpyxl
        year_id = self.year_box.currentData()
        term_id = self.term_box.currentData()
        class_name = self.class_box.currentText()
        level = SystemState.get_level()

        if not self.enrollment_mode_checkbox.isChecked():
            show_error(self, "Enable Enrollment Mode before importing enrollments.")
            return

        if not (year_id and term_id and class_name and class_name != "-- Select Class --"):
            show_error(self, "Select Year, Term, and Class before importing enrollments.")
            return

        path = excel_utils.get_import_file(self)
        if not path:
            return

        progress = excel_utils.make_progress(self, "Importing enrollments...")
        try:
            excel_utils.set_progress(progress, 5, "Reading workbook")
            wb = openpyxl.load_workbook(path, data_only=True)
            sheet = wb.active

            students = {
                str(admission).strip(): str(name)
                for admission, name in fetch_all(
                    """
                    SELECT admission_no, full_name
                    FROM students
                    WHERE class=? AND level=?
                    """,
                    (class_name, level),
                )
            }
            subjects = {
                str(subject).strip().casefold(): str(subject).strip()
                for (subject,) in fetch_all(
                    "SELECT subject_name FROM subjects WHERE level=?",
                    (level,),
                )
            }
            existing = {
                (str(admission).strip(), str(subject).strip())
                for admission, subject in fetch_all(
                    """
                    SELECT admission_no, subject_name
                    FROM enrollments
                    WHERE academic_year_id=? AND term_id=? AND class_name=?
                    """,
                    (year_id, term_id, class_name),
                )
            }
            excel_utils.set_progress(progress, 25, "Detecting template format")

            # Locate the header row structurally (it shifts depending on how
            # many instruction lines the template was generated with) rather
            # than assuming a fixed row number. The legacy long-format
            # template has exactly ["Admission No*", "Subject Name*"]; the
            # wide-format template has one column per subject name.
            header_row_num = excel_utils.find_header_row(sheet)
            if header_row_num is None:
                raise ValueError(
                    "Could not detect the template's header row. Please use a "
                    "template downloaded from this system, and don't remove "
                    "the 'INSTRUCTIONS:' block or reorder rows."
                )
            data_start_row = header_row_num + 2

            header_row = next(sheet.iter_rows(min_row=header_row_num, max_row=header_row_num, values_only=True), ())
            header_values = [str(h).strip() for h in header_row if h not in (None, "")]
            is_long_format = header_values[:2] == ["Admission No*", "Subject Name*"]

            excel_utils.set_progress(progress, 40, "Reading rows")
            if is_long_format:
                candidate_pairs = self._read_long_format_pairs(sheet, data_start_row)
            else:
                candidate_pairs = self._read_wide_format_pairs(sheet, header_row, data_start_row)
            excel_utils.set_progress(progress, 60, "Validating entries")

            new_enrollments = []
            seen = set()
            invalid_students = []
            invalid_subjects = []
            duplicate_rows = 0
            existing_rows = 0

            for row_label, admission_raw, subject_raw in candidate_pairs:
                admission_no = str(admission_raw).strip()
                subject_key = str(subject_raw).strip().casefold()
                subject_name = subjects.get(subject_key)

                if admission_no not in students:
                    invalid_students.append(f"{row_label}: {admission_no}")
                    continue
                if not subject_name:
                    invalid_subjects.append(f"{row_label}: {subject_raw}")
                    continue

                enrollment = (admission_no, subject_name)
                if enrollment in seen:
                    duplicate_rows += 1
                    continue
                seen.add(enrollment)

                if enrollment in existing:
                    existing_rows += 1
                    continue
                new_enrollments.append(enrollment)

            excel_utils.set_progress(progress, 80, "Preparing summary")
            summary = [
                f"New enrollments to add: {len(new_enrollments)}",
                f"Existing enrollments kept unchanged: {existing_rows}",
                f"Duplicate spreadsheet rows skipped: {duplicate_rows}",
                f"Invalid student rows: {len(invalid_students)}",
                f"Invalid subject rows: {len(invalid_subjects)}",
            ]
            problems = invalid_students[:3] + invalid_subjects[:3]
            if problems:
                summary.append("\nExamples requiring correction:\n" + "\n".join(problems))

            progress.close()

            if not new_enrollments:
                show_info(self, "No new enrollments were found.\n\n" + "\n".join(summary), title="Import Preview")
                return

            reply = QMessageBox.question(
                self,
                "Confirm Enrollment Import",
                "\n".join(summary) + "\n\nAdd the valid new enrollments?",
                QMessageBox.Yes | QMessageBox.No,
            )
            if reply != QMessageBox.Yes:
                return

            progress = excel_utils.make_progress(self, "Saving enrollments...")
            with get_cursor(commit=True) as cur:
                cur.executemany(
                    """
                    INSERT INTO enrollments (admission_no, subject_name, class_name, academic_year_id, term_id)
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    [
                        (admission_no, subject_name, class_name, year_id, term_id)
                        for admission_no, subject_name in new_enrollments
                    ],
                )
            excel_utils.set_progress(progress, 100, "Saved")
            progress.close()

            self.load_enrollment_data()
            show_info(
                self,
                f"Imported {len(new_enrollments)} new enrollment record(s).\n\n"
                + "\n".join(summary[1:]),
                title="Import Complete",
            )

        except Exception as e:
            logger.exception("Enrollment import failed for file=%s class=%s", path, class_name)
            QMessageBox.critical(self, "Error", f"Import failed: {e}")
        finally:
            if progress is not None:
                progress.close()