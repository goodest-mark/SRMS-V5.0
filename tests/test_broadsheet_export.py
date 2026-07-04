import os
from pathlib import Path

import openpyxl
from PySide6.QtWidgets import QApplication, QFileDialog, QMessageBox

from broadsheet_export import to_excel


def test_broadsheet_excel_headers_use_short_names(monkeypatch, tmp_path):
    app = QApplication.instance() or QApplication([])

    output_path = tmp_path / "broadsheet.xlsx"
    monkeypatch.setattr(QFileDialog, "getSaveFileName", lambda parent, title, fname, filter: (str(output_path), filter))
    monkeypatch.setattr(QMessageBox, "information", lambda parent, title, msg: None)
    monkeypatch.setattr(QMessageBox, "critical", lambda parent, title, msg: None)

    data = {
        'subjects': ['Mathematics', 'English Language'],
        'subject_headers': ['MATH', 'ENGL'],
        'rows': [
            {
                'Position': 1,
                'Admission No': 'ADM001',
                'Student Name': 'Jane Doe',
                'Gender': 'Female',
                'marks': {'Mathematics': 85, 'English Language': 92},
                'Total': 177,
                'Average': 88.5,
                'Points': 8,
                'Division': 'I'
            }
        ],
        'meta': {
            'class': 'Form 1A',
            'level': 'O_LEVEL',
            'exam': 'Midterm',
            'term': 'Term 1',
            'year': '2025',
            'school_profile': {
                'school_name': 'Test School',
                'school_address': '123 Test Road',
                'school_phone': '123456',
                'school_email': 'info@test.edu'
            }
        },
        'class_performance': {'total_students': 1, 'class_average': 88.5, 'highest_average': 88.5, 'lowest_average': 88.5, 'pass_count': 1, 'fail_count': 0, 'pass_rate': 100, 'fail_rate': 0},
        'gender_summary': {'Male': 0, 'Female': 1, 'Total': 1},
        'division_summary': {'I': 1, 'II': 0, 'III': 0, 'IV': 0, '0': 0, 'Incomplete': 0},
        'top_students': [],
        'bottom_students': [],
        'subject_performance': {'Mathematics': {'average': 85, 'passes': 1, 'fails': 0}, 'English Language': {'average': 92, 'passes': 1, 'fails': 0}},
        'subject_ranking': [('English Language', {'average': 92}), ('Mathematics', {'average': 85})],
        'best_subject': 'English Language',
        'worst_subject': 'Mathematics',
        'max_avg': 92,
        'min_avg': 85,
        'settings': {'show_gender_summary': True, 'show_subject_ranking': True, 'show_logo': False, 'show_watermark': False}
    }

    to_excel(None, data)

    assert output_path.exists()
    workbook = openpyxl.load_workbook(output_path)
    ws = workbook.active
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if row and row[0] == 'Position' and 'Admission No' in row:
            header_row = list(row)
            break

    assert header_row == ['Position', 'Admission No', 'Student Name', 'Gender', 'MATH', 'ENGL', 'Total', 'Average', 'Points', 'Division']
