import os
import random
import shutil
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SEED = 20260705


def create_styled_excel(filename, title, headers, data, sample=None, instructions=None):
    """Generate an Excel file that matches the SRMS import template style."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"

    blue_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1, value="STATE UNIVERSITY DEMONSTRATION SECONDARY SCHOOL").font = Font(size=16, bold=True)
    ws.cell(row=1, column=1).alignment = center_align

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.cell(row=2, column=1, value="Tanzania | srms@example.com").font = Font(size=10)
    ws.cell(row=2, column=1).alignment = center_align

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(headers))
    ws.cell(row=3, column=1, value=title.upper()).font = Font(size=14, bold=True, color="2563EB")
    ws.cell(row=3, column=1).alignment = center_align

    ws.cell(row=5, column=1, value="INSTRUCTIONS:").font = bold_font
    if instructions is None:
        instructions = [
            "1. Do not modify the column headers in Row 10.",
            "2. Start data entry from Row 12 (below the sample row).",
            "3. Keep duplicate rows if you want to test overwrite or conflict handling.",
            "4. Leave a row in place even if it is invalid when you want to test rejection paths.",
        ]
    for i, text in enumerate(instructions):
        ws.cell(row=6 + i, column=1, value=text).font = Font(italic=True, size=9)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col_num, value=header)
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = center_align

    if sample:
        for col_num, val in enumerate(sample, 1):
            cell = ws.cell(row=11, column=col_num, value=val)
            cell.font = Font(italic=True, color="808080")

    for r_idx, row_data in enumerate(data, 12):
        for c_idx, value in enumerate(row_data, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws.freeze_panes = "A11"

    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 4

    wb.save(filename)


def write_manifest(pack_dir, files):
    manifest = Path(pack_dir) / "PACK_MANIFEST.md"
    lines = [
        "# SRMS V5 Test Pack",
        "",
        "This pack is intentionally mixed: valid rows, duplicates, boundary values, and invalid rows are included so the import paths can be exercised from multiple angles.",
        "",
        "## Suggested order",
        "",
        "1. Import students.",
        "2. Import subjects.",
        "3. Import enrollments.",
        "4. Import requirements.",
        "5. Import results after the relevant enrollments exist.",
        "",
        "## Files",
        "",
    ]
    for name, note in files:
        lines.append(f"- `{name}` - {note}")
    manifest.write_text("\n".join(lines) + "\n", encoding="utf-8")


def build_students(rng):
    o_students = [
        ["2024/0001", "EX/2024/0001", "Asha Mzava", "Female", "Form I", "A", "O_LEVEL", "Baseline valid row"],
        ["2024/0002", "EX/2024/0002", "Juma Mollel", "Male", "Form I", "B", "O_LEVEL", "Baseline valid row"],
        ["2024/0003", "EX/2024/0003", "Neema Ally", "Female", "Form II", "A", "O_LEVEL", "Baseline valid row"],
        ["2024/0004", "EX/2024/0004", "Hassan Said", "Male", "Form II", "B", "O_LEVEL", "Baseline valid row"],
        ["2024/0005", "EX/2024/0005", "Fatuma Juma", "Female", "Form III", "A", "O_LEVEL", "Baseline valid row"],
        ["2024/0006", "EX/2024/0006", "Peter John", "Male", "Form III", "C", "O_LEVEL", "Baseline valid row"],
        ["2024/0007", "EX/2024/0007", "Grace Mtega", "Female", "Form IV", "B", "O_LEVEL", "Baseline valid row"],
        ["2024/0008", "EX/2024/0008", "Salum Kassim", "Male", "Form IV", "C", "O_LEVEL", "Baseline valid row"],
        ["2024/0001", "EX/2024/9001", "Asha Mzava Updated", "Female", "Form I", "B", "O_LEVEL", "Duplicate admission should update"],
        ["2024/0009", "EX/2024/0009", "", "Female", "Form I", "A", "O_LEVEL", "Missing name should be rejected"],
        ["2024/0010", "EX/2024/0010", "Invalid Class Student", "Male", "Form VII", "A", "O_LEVEL", "Invalid class should be rejected"],
        ["2024/0011", "EX/2024/0011", "Annah   Mhando", "Female", "Form II", "", "O_LEVEL", "Stream is optional"],
        ["2024/0012", "EX/2024/0012", "Rashid Khamis", "Other", "Form III", "A", "O_LEVEL", "Non-standard gender string"],
        ["2024/0013", "EX/2024/0013", "Emmanuel Lema", "Male", "Form IV", "B", "O_LEVEL", "Valid row with punctuation !@#"],
        ["2024/0014", "", "Whitespace Name", "Female", "Form I", "A", "O_LEVEL", "Exam number blank is allowed"],
    ]

    a_students = [
        ["2024/0101", "EX/2024/0101", "Maryam Said", "Female", "Form V", "A", "A_LEVEL", "Baseline valid row"],
        ["2024/0102", "EX/2024/0102", "Emmanuel Mchengerwa", "Male", "Form V", "B", "A_LEVEL", "Baseline valid row"],
        ["2024/0103", "EX/2024/0103", "Amina Juma", "Female", "Form VI", "A", "A_LEVEL", "Baseline valid row"],
        ["2024/0104", "EX/2024/0104", "Godfrey Lema", "Male", "Form VI", "B", "A_LEVEL", "Baseline valid row"],
        ["2024/0105", "EX/2024/0105", "Amina Juma", "Female", "Form VI", "B", "A_LEVEL", "Duplicate name different admission"],
        ["2024/0106", "EX/2024/0106", "   Alice   Mtei   ", "Female", "Form V", "", "A_LEVEL", "Whitespace handling"],
        ["2024/0107", "EX/2024/0107", "Long Name " + "Test " * 8, "Male", "Form V", "C", "A_LEVEL", "Long name edge case"],
        ["2024/0108", "EX/2024/0108", "Missing Class Student", "Female", "", "A", "A_LEVEL", "Missing class should be rejected"],
    ]

    legacy_students = [
        ["2024/0201", "Asha Legacy", "Female", "Form I", "A", "O_LEVEL", "Legacy 7-column layout"],
        ["2024/0201", "Asha Legacy Updated", "Female", "Form I", "B", "O_LEVEL", "Duplicate admission in legacy file"],
        ["2024/0202", "Juma Legacy", "Male", "Form II", "A", "O_LEVEL", "Legacy row"],
        ["2024/0203", "", "Male", "Form III", "A", "O_LEVEL", "Missing name should be rejected"],
        ["2024/0204", "Invalid Legacy Class", "Male", "Form VII", "A", "O_LEVEL", "Invalid class should be rejected"],
        ["2024/0205", "Maryam Legacy", "Female", "Form V", "B", "A_LEVEL", "A-Level legacy row"],
    ]

    return o_students, a_students, legacy_students


def build_subjects():
    return [
        ["Mathematics", "COUNTED", "O_LEVEL"],
        ["Mathematics", "PRINCIPAL", "O_LEVEL"],
        ["English", "COUNTED", "O_LEVEL"],
        ["Kiswahili", "COUNTED", "O_LEVEL"],
        ["Biology", "COUNTED", "O_LEVEL"],
        ["Chemistry", "COUNTED", "O_LEVEL"],
        ["Physics", "COUNTED", "O_LEVEL"],
        ["History", "COUNTED", "O_LEVEL"],
        ["Geography", "COUNTED", "O_LEVEL"],
        ["Civics", "COUNTED", "O_LEVEL"],
        ["Agriculture", "COUNTED", "O_LEVEL"],
        ["History", "PRINCIPAL", "A_LEVEL"],
        ["Geography", "PRINCIPAL", "A_LEVEL"],
        ["Economics", "PRINCIPAL", "A_LEVEL"],
        ["General Studies", "SUBSIDIARY", "A_LEVEL"],
        ["General Studies", "PRINCIPAL", "A_LEVEL"],
        ["Mathematics", "COUNTED", "A_LEVEL"],
    ]


def build_enrollments():
    rows = []
    o_students = [f"2024/{index:04d}" for index in range(1, 9)]
    a_students = [f"2024/{index:04d}" for index in range(101, 109)]

    o_subjects = ["Mathematics", "English", "Kiswahili", "Biology", "Chemistry", "Physics", "History", "Geography", "Civics", "Agriculture"]
    a_subjects = ["History", "Geography", "Economics", "General Studies"]

    for adm in o_students:
        for subject in o_subjects[:6]:
            rows.append([adm, subject])
    rows.extend([
        ["2024/0001", "Mathematics"],
        ["2024/0001", "Mathematics"],
        ["2024/0002", "mathematics"],
        ["2024/0003", "Astronomy"],
        ["2024/0999", "English"],
        ["", "Physics"],
        ["2024/0004", ""],
        ["2024/0005", "General Studies"],
    ])

    for adm in a_students:
        for subject in a_subjects:
            rows.append([adm, subject])
    rows.extend([
        ["2024/0101", "History"],
        ["2024/0101", "Economics"],
        ["2024/0102", "General Studies"],
        ["2024/0103", "Mathematics"],
        ["2024/0104", "Geography"],
    ])

    return rows


def build_requirements():
    return [
        ["Reams of Paper", "2", "Urgent"],
        ["Reams of Paper", "4", "Duplicate item to test overwrites in manual workflows"],
        ["Laboratory Contribution", "1", "For practical subjects"],
        ["Examination Fee", "3", "Mid-term collection"],
        ["Identity Cards", "20", "Bulk order"],
        ["Sports Contribution", "1", "Includes jerseys and balls"],
        ["School Development Fund", "5", "Long note with punctuation !@#"],
        ["Transport and Welfare", "0", "Zero quantity edge case"],
        ["Very Long Requirement Item Name Designed To Stretch The Column Width", "12", "Long item name"],
        ["Blank Quantity Item", "", "Quantity intentionally blank"],
        ["Optional Notes Item", "7", ""],
    ]


def build_results_o():
    return [
        ["2024/0001", 95],
        ["2024/0001", 97],
        ["2024/0002", 0],
        ["2024/0003", 100],
        ["2024/0004", 64.8],
        ["2024/0005", -1],
        ["2024/0006", 101],
        ["2024/0007", "A"],
        ["2024/0008", None],
        ["2024/0101", 55],
        ["2024/0999", 77],
        ["", 40],
    ]


def build_results_a():
    return [
        ["2024/0101", 88],
        ["2024/0101", 91],
        ["2024/0102", 73],
        ["2024/0103", 0],
        ["2024/0104", 100],
        ["2024/0105", 52.5],
        ["2024/0106", -3],
        ["2024/0107", 105],
        ["2024/0001", 66],
        ["2024/0999", 44],
        ["", 61],
        ["2024/0108", "Seventy"],
    ]


def generate_pack():
    rng = random.Random(SEED)
    pack_dir = Path("SRMS_V5_TEST_PACK")
    if pack_dir.exists():
        shutil.rmtree(pack_dir)
    pack_dir.mkdir()

    files = []

    o_students, a_students, legacy_students = build_students(rng)
    student_headers = ["Admission No", "Exam No", "Full Name", "Gender", "Class", "Stream", "Level", "Comments"]
    legacy_headers = ["Admission No", "Full Name", "Gender", "Class", "Stream", "Level", "Comments"]

    create_styled_excel(
        pack_dir / "students_import.xlsx",
        "STUDENT REGISTRATION FORM - MIXED CASES",
        student_headers,
        o_students + a_students,
        sample=["2024/0001", "EX/2024/0001", "Asha Mzava", "Female", "Form I", "A", "O_LEVEL", "Valid example"],
        instructions=[
            "1. The first block is O-Level data, the second block is A-Level data.",
            "2. Duplicate admission numbers are intentional and should update existing records.",
            "3. Rows with missing names or invalid classes should be rejected by the import logic.",
            "4. Keep the same sheet structure so the import path sees both valid and invalid rows.",
        ],
    )
    files.append(("students_import.xlsx", "Mixed O-Level and A-Level student rows, plus duplicates and invalid records."))

    create_styled_excel(
        pack_dir / "students_legacy_import.xlsx",
        "STUDENT REGISTRATION FORM - LEGACY LAYOUT",
        legacy_headers,
        legacy_students,
        sample=["2024/0201", "Asha Legacy", "Female", "Form I", "A", "O_LEVEL", "Legacy example"],
        instructions=[
            "1. This file uses the older 7-column student layout supported by the importer.",
            "2. It includes duplicate admissions, missing names, and invalid classes for rejection testing.",
        ],
    )
    files.append(("students_legacy_import.xlsx", "Legacy 7-column student layout for backward-compatibility testing."))

    subjects = build_subjects()
    create_styled_excel(
        pack_dir / "subjects_import.xlsx",
        "SUBJECT REGISTRATION FORM - MIXED CASES",
        ["Subject Name", "Subject Type", "Level"],
        subjects,
        sample=["Mathematics", "COUNTED", "O_LEVEL"],
        instructions=[
            "1. Duplicate subject names are intentional so the conflict-update branch can be exercised.",
            "2. Include both O-Level and A-Level subjects in one file to test level-scoped imports.",
            "3. A repeated subject with a different type should update the existing row.",
        ],
    )
    files.append(("subjects_import.xlsx", "Duplicate subjects, type changes, and cross-level subject coverage."))

    enrollments = build_enrollments()
    create_styled_excel(
        pack_dir / "enrollments_import.xlsx",
        "STUDENT SUBJECT ENROLLMENT FORM - MIXED CASES",
        ["Admission No", "Subject Name"],
        enrollments,
        sample=["2024/0001", "Mathematics"],
        instructions=[
            "1. Rows include duplicate admissions, case mismatches, unknown students, and unknown subjects.",
            "2. Import the file once for an O-Level class and again for an A-Level class to exercise both branches.",
            "3. Rows that do not map to an existing student/subject pair should be ignored by the importer.",
        ],
    )
    files.append(("enrollments_import.xlsx", "Enrollment stress file with duplicates, cross-level rows, blanks, and unknown references."))

    requirements = build_requirements()
    create_styled_excel(
        pack_dir / "requirements_import.xlsx",
        "SCHOOL REQUIREMENTS TEMPLATE - EDGE CASES",
        ["Item Name", "Quantity", "Notes"],
        requirements,
        sample=["Reams of Paper", "2", "Urgent"],
        instructions=[
            "1. Duplicate items are intentional.",
            "2. Blank quantity and very long item names are included to test resilience.",
            "3. Notes may be empty or contain punctuation.",
        ],
    )
    files.append(("requirements_import.xlsx", "Requirement rows with duplicates, blank quantities, long names, and punctuation."))

    o_result_rows = build_results_o()
    a_result_rows = build_results_a()
    o_subjects = ["Mathematics", "English", "Kiswahili", "Biology", "Chemistry", "Physics", "History", "Geography", "Civics"]
    a_subjects = ["History", "Geography", "Economics", "General Studies"]

    for subject in o_subjects:
        file_name = f"{subject.lower()}_results.xlsx"
        create_styled_excel(
            pack_dir / file_name,
            f"{subject.upper()} RESULTS - EDGE CASES",
            ["Admission No", "Marks"],
            o_result_rows,
            sample=["2024/0001", 95],
            instructions=[
                f"1. Select {subject} and an O-Level class before importing this file.",
                "2. The file includes duplicate admissions, invalid marks, missing admissions, and a non-enrolled student.",
                "3. Valid rows should save, while invalid rows should be rejected or skipped by the importer.",
            ],
        )
        files.append((file_name, f"O-Level results for {subject} with duplicate marks and invalid rows."))

    for subject in a_subjects:
        file_name = f"{subject.lower().replace(' ', '_')}_al_results.xlsx"
        create_styled_excel(
            pack_dir / file_name,
            f"{subject.upper()} RESULTS - EDGE CASES",
            ["Admission No", "Marks"],
            a_result_rows,
            sample=["2024/0101", 88],
            instructions=[
                f"1. Select {subject} and an A-Level class before importing this file.",
                "2. The file includes duplicate admissions, invalid marks, missing admissions, and a non-enrolled O-Level student.",
                "3. Valid rows should save, while invalid rows should be rejected or skipped by the importer.",
            ],
        )
        files.append((file_name, f"A-Level results for {subject} with duplicate marks and invalid rows."))

    write_manifest(pack_dir, files)
    shutil.make_archive(str(pack_dir), "zip", pack_dir)
    print(f"Generated {pack_dir}.zip successfully.")


if __name__ == "__main__":
    generate_pack()
