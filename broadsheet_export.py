import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as ExcelImage
from PySide6.QtWidgets import QFileDialog, QMessageBox, QProgressDialog, QApplication
from reportlab.lib import colors
from ui_helpers import get_subject_short_name
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.platypus import PageBreak, Frame, PageTemplate
from reportlab.platypus.flowables import Image
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from PySide6.QtCore import Qt
from settings_page import get_setting
from grade_utils import get_grade

# ------------------------------------------------------------
# Helper functions
# ------------------------------------------------------------

def _make_progress(parent, title):
    if QApplication.instance() is None:
        return None
    progress = QProgressDialog(title, None, 0, 100, parent)
    progress.setWindowTitle(title)
    progress.setWindowModality(Qt.WindowModal if parent else Qt.NonModal)
    progress.setAutoClose(False)
    progress.setAutoReset(False)
    progress.setMinimumDuration(0)
    progress.setValue(0)
    QApplication.processEvents()
    return progress

def _set_progress(progress, value, label):
    if progress is None:
        return
    progress.setLabelText(f"{label}\n\nProgress: {value}%")
    progress.setValue(value)
    QApplication.processEvents()

def _report_palette():
    return {
        "accent": "2563EB",
        "accent_dark": "1B3A5C",
        "light": "E8EDF2",
        "text": "111827",
    }

def _hex_color(value):
    return colors.HexColor(f"#{value}")

def _pick(mapping, *keys, default="-"):
    for key in keys:
        if isinstance(mapping, dict) and key in mapping:
            return mapping[key]
    return default

def _format_gender(gender):
    if gender:
        g = str(gender).lower().strip()
        if g.startswith('m'):
            return "M"
        elif g.startswith('f'):
            return "F"
    return gender or "-"

def _filter_active_subjects(rows, subjects):
    """
    Keep only subjects that at least one student actually has a mark for.
    Subjects where every row is None/'-'/'' (nobody sat the exam) are dropped
    so they don't waste columns / rows in the report.
    """
    active = []
    for s in subjects:
        for r in rows:
            mark = r['marks'].get(s)
            if mark is not None and mark != '-' and mark != '':
                active.append(s)
                break
    return active

def _get_grade_point(grade, level):
    """
    Return the grade point for a given grade and level.
    O-Level: A=1, B=2, C=3, D=4, F=5
    A-Level: A=1, B=2, C=3, D=4, E=5, S=6, F=7
    Returns None if grade is not recognised.
    """
    if level == "O_LEVEL":
        mapping = {'A': 1, 'B': 2, 'C': 3, 'D': 4, 'F': 5}
    else:  # A_LEVEL
        mapping = {'A': 1, 'B': 2, 'C': 3, 'D': 4, 'E': 5, 'S': 6, 'F': 7}
    return mapping.get(grade)  # Returns None if grade not found

def _compute_subject_gpa(rows, subject, level):
    """
    Compute grade‑based GPA for a subject:
    sum of grade points ÷ number of students who sat.
    Returns float (rounded to 2 decimals) or "N/A" if no valid grades.
    """
    total_points = 0
    count = 0
    for r in rows:
        mark = r['marks'].get(subject)
        if mark is not None and mark != '-' and mark != '':
            try:
                mark_val = float(mark)
                grade = get_grade(mark_val, level=level)
                if grade is None:
                    continue
                point = _get_grade_point(grade, level)
                if point is not None:
                    total_points += point
                    count += 1
                else:
                    print(f"[WARNING] Unknown grade '{grade}' for subject '{subject}' (skipping)")
            except (ValueError, TypeError) as e:
                print(f"[WARNING] Could not convert mark '{mark}' for subject '{subject}': {e}")
                continue
    if count == 0:
        return "N/A"
    return round(total_points / count, 2)

def _compute_subject_mean_mark(rows, subject):
    """
    Compute raw‑marks average for a subject:
    sum of marks ÷ number of students who sat.
    Returns float (rounded to 2 decimals) or "N/A" if no valid marks.
    """
    total = 0
    count = 0
    for r in rows:
        mark = r['marks'].get(subject)
        if mark is not None and mark != '-' and mark != '':
            try:
                total += float(mark)
                count += 1
            except (ValueError, TypeError):
                continue
    if count == 0:
        return "N/A"
    return round(total / count, 2)

# ------------------------------------------------------------
# Excel export
# ------------------------------------------------------------

def to_excel(parent, data):
    path, _ = QFileDialog.getSaveFileName(parent, "Export Broadsheet", f"Broadsheet_{data['meta']['class']}.xlsx", "Excel Files (*.xlsx)")
    if not path:
        return

    progress = None
    try:
        progress = _make_progress(parent, "Exporting Excel broadsheet...")
        wb = openpyxl.Workbook()
        _set_progress(progress, 5, "Preparing workbook")
        ws = wb.active
        ws.title = "Broadsheet"

        meta = data['meta']
        school_profile = meta['school_profile']
        palette = _report_palette()
        subjects = _filter_active_subjects(data['rows'], data['subjects'])
        last_col = len(subjects) + 8
        level = meta['level']

        school_name = school_profile.get('school_name', 'SCHOOL MANAGEMENT SYSTEM').upper()
        motto = school_profile.get('school_motto') or school_profile.get('motto') or ''
        address = school_profile.get('school_address') or school_profile.get('address', '-') or '-'
        phone = school_profile.get('school_phone') or school_profile.get('phone') or '-'
        email = school_profile.get('school_email') or school_profile.get('email') or '-'
        website = school_profile.get('school_website') or ''
        logo_path = school_profile.get('school_logo')

        if get_setting('show_logo', '1') == '1' and logo_path and os.path.exists(logo_path):
            try:
                img = ExcelImage(logo_path)
                img.width = 100
                img.height = 100
                ws.add_image(img, 'A1')
                start_row = 1
                start_col = 3
            except Exception as e:
                print(f"[WARNING] Could not add logo to Excel: {e}")
                start_row = 1
                start_col = 1
        else:
            start_row = 1
            start_col = 1

        header_rows = [
            (school_name, 18, palette['accent_dark'], "FFFFFF"),
            (motto, 11, palette['accent_dark'], "FFFFFF"),
            (f"{address} | Tel: {phone} | Email: {email} | {website}", 10, palette['light'], palette['text']),
            ("CLASS BROADSHEET / REPORT SUMMARY", 14, palette['accent'], "FFFFFF"),
            (f"Class: {meta['class']}   Level: {meta['level']}   Exam: {meta['exam']}   Term: {meta['term']}   Year: {meta['year']}", 11, palette['light'], palette['text']),
        ]

        for idx, (text, size, fill, font_color) in enumerate(header_rows):
            row_no = start_row + idx
            ws.merge_cells(start_row=row_no, start_column=start_col, end_row=row_no, end_column=last_col)
            cell = ws.cell(row=row_no, column=start_col)
            cell.value = text
            cell.font = Font(size=size, bold=True, color=font_color)
            cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[row_no].height = 24 if size >= 14 else 19

        subject_headers = [get_subject_short_name(s) for s in subjects]
        headers = ["Position", "Admission No", "Student Name", "Gender"] + subject_headers + ["Total", "Average", "Points", "Division"]

        ws.append([])
        ws.append(headers)
        header_row = ws.max_row
        for cell in ws[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=palette["accent_dark"], end_color=palette["accent_dark"], fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        rows_data = data['rows']
        for row_index, r in enumerate(rows_data, start=1):
            gender_display = _format_gender(r['Gender'])
            row_vals = [r['Position'], r['Admission No'], r['Student Name'], gender_display]
            for s in subjects:
                row_vals.append(r['marks'].get(s, "-"))
            row_vals += [r['Total'], r['Average'], r['Points'], r['Division']]
            ws.append(row_vals)
            percent = 10 + int((row_index / max(len(rows_data), 1)) * 70)
            _set_progress(progress, percent, f"Writing student {row_index}/{len(rows_data)}")

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for r_idx in range(header_row, ws.max_row + 1):
            for c_idx in range(1, len(headers) + 1):
                ws.cell(row=r_idx, column=c_idx).border = thin_border

        ws.append([])

        # ----- Analytics: Class Performance -----
        class_perf = data['class_performance']
        ws.append(["CLASS PERFORMANCE ANALYSIS"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.append(["Total Students", class_perf['total_students'], "Class Average", f"{class_perf['class_average']}%"])
        ws.append(["Highest Average", f"{class_perf['highest_average']}%", "Lowest Average", f"{class_perf['lowest_average']}%"])
        ws.append(["Pass Rate", f"{class_perf['pass_rate']}%", "Fail Rate", f"{class_perf['fail_rate']}%"])
        ws.append([])

        # ----- Analytics: Gender Summary -----
        if data['settings']['show_gender_summary']:
            ws.append(["GENDER SUMMARY"])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
            gender_sum = data['gender_summary']
            ws.append(["Gender", "Count"])
            ws.append(["M", gender_sum.get('Male', 0)])
            ws.append(["F", gender_sum.get('Female', 0)])
            ws.append(["Total", gender_sum['Total']])
            ws.append([])

        # ----- Analytics: Division Summary -----
        ws.append(["DIVISION SUMMARY"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.append(["Division", "Students"])
        for div, count in data['division_summary'].items():
            ws.append([div, count])
        ws.append([])

        # ----- Analytics: Top 10 Students -----
        ws.append(["TOP 10 STUDENTS"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.append(["Position", "Admission No", "Student Name", "Average", "Division"])
        for s in data['top_students']:
            ws.append([s['position'], s['admission'], s['name'], s['average'], s['division']])
        ws.append([])

        # ----- Analytics: Bottom 10 Students -----
        ws.append(["BOTTOM 10 STUDENTS"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.append(["Position", "Admission No", "Student Name", "Average", "Division"])
        for s in data['bottom_students']:
            ws.append([s['position'], s['admission'], s['name'], s['average'], s['division']])
        ws.append([])

        # ----- Subject Performance Analysis (with GPA) -----
        ws.append(["SUBJECT PERFORMANCE ANALYSIS"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        # Headers: Subject, Average (mark), Passes, Fails, GPA (grade‑based)
        ws.append(["Subject", "Average", "Passes", "Fails", "GPA"])

        best_sub = None
        worst_sub = None
        min_gpa = 999
        max_gpa = -1
        for sub_name, stats in data['subject_performance'].items():
            if sub_name not in subjects:
                continue
            display_name = get_subject_short_name(sub_name)
            # Compute GPA using grade points (level‑aware)
            gpa = _compute_subject_gpa(rows_data, sub_name, level)
            # Keep existing raw‑marks average, passes, fails
            avg_mark = stats['average']
            passes = stats['passes']
            fails = stats['fails']
            ws.append([display_name, avg_mark, passes, fails, gpa])

            # Track best/worst for summary (only if GPA is a number)
            if isinstance(gpa, (int, float)):
                if gpa < min_gpa or min_gpa == 999:
                    min_gpa = gpa
                    best_sub = display_name
                if gpa > max_gpa or max_gpa == -1:
                    max_gpa = gpa
                    worst_sub = display_name

        ws.append([f"Best Subject (GPA): {best_sub} (GPA: {min_gpa})"])
        ws.append([f"Worst Subject (GPA): {worst_sub} (GPA: {max_gpa})"])
        ws.append([])

        # ----- Signatures -----
        ws.append([])
        ws.append(["Academic Master / Mistress Signature: ......................................."])
        ws.append(["Headmaster / Headmistress Signature: ........................................"])
        ws.append(["School Stamp:"])

        # Style data cells
        for row in ws.iter_rows(min_row=header_row + 1):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

        ws.freeze_panes = f"E{header_row + 1}"

        _set_progress(progress, 90, "Saving Excel file")
        wb.save(path)
        _set_progress(progress, 100, "Export completed")
        QMessageBox.information(parent, "Success", f"Broadsheet exported to {path}")
    except Exception as e:
        print(f"[ERROR] Broadsheet export failed: {e}")
        QMessageBox.critical(parent, "Export Error", "An unexpected error occurred during export.")
    finally:
        if progress is not None:
            progress.close()

# ------------------------------------------------------------
# PDF export
# ------------------------------------------------------------

def to_pdf(parent, data):
    path, _ = QFileDialog.getSaveFileName(parent, "Export Broadsheet PDF", f"Broadsheet_{data['meta']['class']}.pdf", "PDF Files (*.pdf)")
    if not path:
        return

    progress = None
    try:
        progress = _make_progress(parent, "Exporting PDF broadsheet...")
        doc = SimpleDocTemplate(path, pagesize=landscape(A4),
                                rightMargin=0.5*inch, leftMargin=0.5*inch,
                                topMargin=0.75*inch, bottomMargin=0.5*inch)

        _set_progress(progress, 8, "Preparing PDF document")
        styles = getSampleStyleSheet()

        # Custom styles – compact and clear
        muted_white = colors.HexColor('#B8C7DB')
        styles.add(ParagraphStyle(name='BrandTitle', alignment=TA_CENTER, fontSize=21, fontName='Helvetica-Bold', leading=23, textColor=colors.white))
        styles.add(ParagraphStyle(name='BrandMotto', alignment=TA_CENTER, fontSize=9, fontName='Helvetica-Oblique', leading=10, textColor=muted_white))
        styles.add(ParagraphStyle(name='BrandContact', alignment=TA_CENTER, fontSize=8, fontName='Helvetica', leading=10, textColor=muted_white))
        styles.add(ParagraphStyle(name='BrandSmallLeft', alignment=TA_LEFT, fontSize=8.5, fontName='Helvetica', leading=11, textColor=colors.white))
        styles.add(ParagraphStyle(name='BrandSmallLeftSub', alignment=TA_LEFT, fontSize=8, fontName='Helvetica', leading=10, textColor=muted_white))
        styles.add(ParagraphStyle(name='BrandSmallRight', alignment=TA_RIGHT, fontSize=8.5, fontName='Helvetica', leading=11, textColor=colors.white))
        styles.add(ParagraphStyle(name='BrandSmallRightSub', alignment=TA_RIGHT, fontSize=8, fontName='Helvetica', leading=10, textColor=muted_white))
        styles.add(ParagraphStyle(name='SectionHeader', alignment=TA_LEFT, fontSize=11, fontName='Helvetica-Bold', leading=13, spaceBefore=10, spaceAfter=5))
        styles.add(ParagraphStyle(name='SigHeader', alignment=TA_CENTER, fontSize=9, fontName='Helvetica-Bold', leading=11))

        meta = data['meta']
        school_profile = meta['school_profile']
        settings = data['settings']
        generated_date = meta['generated_date']
        palette = _report_palette()
        accent_dark = _hex_color(palette['accent_dark'])
        accent = _hex_color(palette['accent'])
        level = meta['level']

        def _header_footer(canvas, doc):
            canvas.saveState()
            header_frame = Frame(
                doc.leftMargin,
                doc.height + doc.topMargin - 1.5 * inch,
                doc.width,
                1.5 * inch,
                leftPadding=0,
                bottomPadding=0,
                rightPadding=0,
                topPadding=0,
                showBoundary=0,
            )

            school_name = school_profile.get('school_name', 'SCHOOL MANAGEMENT SYSTEM').upper()
            motto = school_profile.get('school_motto') or ''
            address = school_profile.get('school_address') or '-'
            phone = school_profile.get('school_phone') or '-'
            email = school_profile.get('school_email') or '-'
            website = school_profile.get('school_website') or ''

            # Outer header cells carry 8pt padding on each side (16pt total);
            # the nested tables below must be sized to fit inside that, or
            # their content overflows past the page's printable edge.
            cell_pad = 16
            left_w = doc.width * 0.18 - cell_pad
            center_w = doc.width * 0.50 - cell_pad
            right_w = doc.width * 0.32 - cell_pad

            center_parts = []
            if settings['show_logo'] and school_profile.get('school_logo') and os.path.exists(school_profile['school_logo']):
                try:
                    logo_size = 0.68 * inch
                    logo = Image(school_profile['school_logo'], width=logo_size, height=logo_size)
                    logo_plate = Table([[logo]], colWidths=[logo_size + 8])
                    logo_plate.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, -1), colors.white),
                        ('BOX', (0, 0), (-1, -1), 1, accent),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('TOPPADDING', (0, 0), (-1, -1), 4),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                        ('LEFTPADDING', (0, 0), (-1, -1), 4),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                    ]))
                    center_parts.append([logo_plate])
                except Exception as e:
                    print(f"[WARNING] Could not load PDF logo: {e}")

            center_parts.append([Paragraph(school_name, styles['BrandTitle'])])
            if motto:
                center_parts.append([Paragraph(motto, styles['BrandMotto'])])
            contact_info = f"{address} | Tel: {phone} | Email: {email} | {website}".strip(' |')
            center_parts.append([Paragraph(contact_info, styles['BrandContact'])])

            center_table = Table(center_parts, colWidths=[center_w])
            center_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0.5),
                ('TOPPADDING', (0, 0), (-1, -1), 0.5),
            ]))

            left_info = [
                [Paragraph('<b>REPORT TYPE</b>', styles['BrandSmallLeft'])],
                [Paragraph('CLASS BROADSHEET', styles['BrandSmallLeftSub'])],
                [Paragraph('Academic Summary', styles['BrandSmallLeftSub'])],
            ]
            left = Table(left_info, colWidths=[left_w])
            left.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ]))

            right_info = [
                [Paragraph('<b>ACADEMIC CONTEXT</b>', styles['BrandSmallRight'])],
                [Paragraph(f"Class: {meta['class']} ({meta['level']})", styles['BrandSmallRightSub'])],
                [Paragraph(f"{meta['exam']} | {meta['term']} - {meta['year']}", styles['BrandSmallRightSub'])],
            ]
            right = Table(right_info, colWidths=[right_w])
            right.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ]))

            header = Table(
                [[left, center_table, right]],
                colWidths=[doc.width * 0.18, doc.width * 0.50, doc.width * 0.32],
            )
            header.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), accent_dark),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ('LINEBELOW', (0, 0), (-1, -1), 2.5, accent),
            ]))

            header_frame.addFromList([header], canvas)

            # Divider lines between the three header zones — kept bright enough
            # (mid-tone accent, not a near-navy shade) to still separate cleanly
            # if this is printed or photocopied in black and white.
            band_top = doc.height + doc.topMargin
            band_bottom = band_top - 1.5 * inch + 0.1 * inch
            divider_x1 = doc.leftMargin + doc.width * 0.18
            divider_x2 = doc.leftMargin + doc.width * 0.68
            canvas.setStrokeColor(accent)
            canvas.setLineWidth(0.75)
            canvas.line(divider_x1, band_bottom, divider_x1, band_top - 0.12 * inch)
            canvas.line(divider_x2, band_bottom, divider_x2, band_top - 0.12 * inch)

            canvas.setFont('Helvetica', 7)
            canvas.setFillColor(colors.HexColor('#64748B'))
            canvas.drawString(doc.leftMargin, 0.3 * inch, f"Generated: {generated_date}")
            canvas.drawString(doc.width + doc.leftMargin - 0.5 * inch, 0.3 * inch, f"Page {doc.page}")
            canvas.restoreState()

        def draw_watermark(canvas, doc):
            if settings['show_watermark']:
                canvas.saveState()
                canvas.setFont('Helvetica-Bold', 60)
                canvas.setFillColor(colors.lightgrey, alpha=0.05)
                canvas.translate(doc.width / 2.0, doc.height / 2.0)
                canvas.rotate(45)
                canvas.drawCentredString(0, 0, school_profile.get('watermark_text', 'CONFIDENTIAL'))
                canvas.restoreState()

        frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height - 1.1 * inch,
                      leftPadding=0, bottomPadding=0, rightPadding=0, topPadding=0,
                      showBoundary=0)
        main_template = PageTemplate(id='main_page', frames=[frame], onPage=_header_footer)
        doc.addPageTemplates([main_template])

        elements = []
        _set_progress(progress, 20, "Building broadsheet table")

        # ----- Main student table -----
        subjects_full = _filter_active_subjects(data['rows'], data['subjects'])
        subject_headers = [get_subject_short_name(s) for s in subjects_full]
        headers = ["Pos", "Adm No", "Name", "Sex"] + subject_headers + ["Tot", "Avg", "Pts", "Div"]

        table_data = [headers]
        rows_data = data['rows']
        for r in rows_data:
            gender_display = _format_gender(r['Gender'])
            row_vals = [str(r['Position']), r['Admission No'], r['Student Name'], gender_display]
            for s in subjects_full:
                row_vals.append(str(r['marks'].get(s, '-')))
            row_vals += [str(r['Total']), str(r['Average']), str(r['Points']), str(r['Division']) if r['Division'] else '-']
            table_data.append(row_vals)

        # Dynamic column widths
        fixed_width = 0.4 + 0.8 + 1.5 + 0.4 + 0.6 + 0.6 + 0.6 + 0.6
        remaining_width = landscape(A4)[0] - 1.0 * inch - fixed_width * inch
        if len(subjects_full) > 0:
            subject_width = max(0.35 * inch, min(0.8 * inch, remaining_width / len(subjects_full)))
        else:
            subject_width = 0.5 * inch

        col_widths = [0.4*inch, 0.8*inch, 1.5*inch, 0.4*inch] + [subject_width] * len(subjects_full) + [0.6*inch, 0.6*inch, 0.6*inch, 0.6*inch]
        total_width = sum(col_widths)
        max_width = landscape(A4)[0] - 1.0 * inch
        if total_width > max_width:
            scale = max_width / total_width
            col_widths = [w * scale for w in col_widths]

        font_size = 8
        if len(subjects_full) > 10:
            font_size = 7
        if len(subjects_full) > 15:
            font_size = 6

        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), accent_dark),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), font_size + 1),
            ('FONTSIZE', (0, 1), (-1, -1), font_size),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 1.5, accent),
        ]))
        elements.append(Spacer(1, 0.1*inch))
        elements.append(Paragraph(f"<b>BROADSHEET FOR {meta['class']} ({meta['level']})</b>", styles['SectionHeader']))
        elements.append(Spacer(1, 0.1*inch))
        elements.append(t)
        elements.append(PageBreak())

        _set_progress(progress, 45, "Building analysis sections")

        # ----- 1. Class Performance Analysis -----
        class_perf = data['class_performance']
        elements.append(Paragraph("CLASS PERFORMANCE ANALYSIS", styles['SectionHeader']))
        class_perf_data = [
            ["Total Students:", class_perf['total_students'], "Class Average:", f"{class_perf['class_average']}%"],
            ["Highest Average:", f"{class_perf['highest_average']}%", "Lowest Average:", f"{class_perf['lowest_average']}%"],
            ["Pass Rate:", f"{class_perf['pass_rate']}%", "Fail Rate:", f"{class_perf['fail_rate']}%"]
        ]
        class_perf_table = Table(class_perf_data, colWidths=[2*inch, 1.5*inch, 2*inch, 1.5*inch])
        class_perf_table.setStyle(TableStyle([
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
            ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
            ('FONTNAME', (2,0), (2,-1), 'Helvetica-Bold'),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
            ('LEFTPADDING', (0,0), (-1,-1), 5),
            ('RIGHTPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
            ('TOPPADDING', (0,0), (-1,-1), 3),
        ]))
        elements.append(class_perf_table)
        elements.append(Spacer(1, 0.2*inch))

        # ----- 2. Gender Summary -----
        if settings['show_gender_summary']:
            gender_sum = data['gender_summary']
            elements.append(Paragraph("GENDER SUMMARY", styles['SectionHeader']))
            gender_data = [["Gender", "Count"], ["M", gender_sum.get('Male', 0)], ["F", gender_sum.get('Female', 0)], ["Total", gender_sum['Total']]]
            gender_table = Table(gender_data, colWidths=[2*inch, 1.5*inch])
            gender_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
            ]))
            elements.append(gender_table)
            elements.append(Spacer(1, 0.2*inch))

        # ----- 3. Division Summary -----
        div_sum = data['division_summary']
        elements.append(Paragraph("DIVISION SUMMARY", styles['SectionHeader']))
        div_data = [["Division", "Students"]]
        for div, count in div_sum.items():
            div_data.append([div, count])
        div_table = Table(div_data, colWidths=[2*inch, 1.5*inch])
        div_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), _hex_color(_report_palette()['light'])),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]))
        elements.append(div_table)
        elements.append(Spacer(1, 0.2*inch))

        # ----- 4. Top 10 Students -----
        top_students = data['top_students']
        elements.append(Paragraph("TOP 10 STUDENTS", styles['SectionHeader']))
        top_data = [["Position", "Admission No", "Student Name", "Average", "Division"]]
        for s in top_students:
            top_data.append([_pick(s, 'position', 'Position'), _pick(s, 'admission', 'Admission No'), _pick(s, 'name', 'Student Name'), _pick(s, 'average', 'Average'), _pick(s, 'division', 'Division')])
        top_table = Table(top_data, colWidths=[0.8*inch, 1.2*inch, 2.5*inch, 1*inch, 1*inch])
        top_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), _hex_color(_report_palette()['light'])),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]))
        elements.append(top_table)
        elements.append(Spacer(1, 0.2*inch))

        # ----- 5. Bottom 10 Students -----
        bottom_students = data['bottom_students']
        elements.append(Paragraph("BOTTOM 10 STUDENTS", styles['SectionHeader']))
        bottom_data = [["Position", "Admission No", "Student Name", "Average", "Division"]]
        for s in bottom_students:
            bottom_data.append([_pick(s, 'position', 'Position'), _pick(s, 'admission', 'Admission No'), _pick(s, 'name', 'Student Name'), _pick(s, 'average', 'Average'), _pick(s, 'division', 'Division')])
        bottom_table = Table(bottom_data, colWidths=[0.8*inch, 1.2*inch, 2.5*inch, 1*inch, 1*inch])
        bottom_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), _hex_color(_report_palette()['light'])),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]))
        elements.append(bottom_table)
        elements.append(Spacer(1, 0.2*inch))

        # ----- 6. Subject Performance Analysis (with GPA) -----
        sub_perf = data['subject_performance']
        elements.append(Paragraph("SUBJECT PERFORMANCE ANALYSIS", styles['SectionHeader']))
        # Headers: Subject, Average (mark), Passes, Fails, GPA (grade‑based)
        sub_perf_data = [["Subject", "Average", "Passes", "Fails", "GPA"]]
        best_sub = None
        worst_sub = None
        min_gpa = 999
        max_gpa = -1
        for sub_name, stats in sub_perf.items():
            if sub_name not in subjects_full:
                continue
            display_name = get_subject_short_name(sub_name)
            gpa = _compute_subject_gpa(rows_data, sub_name, level)
            avg_mark = stats['average']
            passes = stats['passes']
            fails = stats['fails']
            sub_perf_data.append([display_name, avg_mark, passes, fails, gpa])

            if isinstance(gpa, (int, float)):
                if gpa < min_gpa or min_gpa == 999:
                    min_gpa = gpa
                    best_sub = display_name
                if gpa > max_gpa or max_gpa == -1:
                    max_gpa = gpa
                    worst_sub = display_name

        sub_perf_table = Table(sub_perf_data, colWidths=[2.5*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch])
        sub_perf_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), _hex_color(_report_palette()['light'])),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]))
        elements.append(sub_perf_table)
        elements.append(Paragraph(f"Best Subject (GPA): {best_sub} (GPA: {min_gpa})", styles['Normal']))
        elements.append(Paragraph(f"Worst Subject (GPA): {worst_sub} (GPA: {max_gpa})", styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))

        # ----- 7. Subject Ranking (by Average Score) -----
        if settings['show_subject_ranking']:
            sub_ranking = data['subject_ranking']
            elements.append(Paragraph("SUBJECT RANKING (by Average Score)", styles['SectionHeader']))
            sub_ranking_data = [["Rank", "Subject"]]
            rank = 0
            for sub_name, stats in sub_ranking:
                if sub_name not in subjects_full:
                    continue
                rank += 1
                display_name = get_subject_short_name(sub_name)
                sub_ranking_data.append([rank, f"{display_name} (Avg: {stats['average']})"])
            sub_ranking_table = Table(sub_ranking_data, colWidths=[1*inch, 3*inch])
            sub_ranking_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
            ]))
            elements.append(sub_ranking_table)
            elements.append(Spacer(1, 0.2*inch))

        # ----- Signatures -----
        elements.append(Spacer(1, 0.4*inch))
        sig_data = [
            [
                Paragraph("<b>CLASS TEACHER</b>", styles['SigHeader']),
                Paragraph("<b>ACADEMIC MASTER / MISTRESS</b>", styles['SigHeader']),
                Paragraph("<b>HEADMASTER / HEADMISTRESS</b>", styles['SigHeader']),
                Paragraph("<b>OFFICIAL STAMP</b>", styles['SigHeader'])
            ],
            [
                Paragraph("Signature: _________________", styles['Normal']),
                Paragraph("Signature: _________________", styles['Normal']),
                Paragraph("Signature: _________________", styles['Normal']),
                ""
            ],
            [
                Paragraph("Date: ______________________", styles['Normal']),
                Paragraph("Date: ______________________", styles['Normal']),
                Paragraph("Date: ______________________", styles['Normal']),
                ""
            ]
        ]
        sig_table = Table(sig_data, colWidths=[doc.width/4]*4)
        sig_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('LINEABOVE', (0,0), (-1,0), 1, accent_dark),
        ]))
        elements.append(sig_table)

        _set_progress(progress, 80, "Rendering PDF pages")
        doc.build(elements, onFirstPage=draw_watermark, onLaterPages=draw_watermark)
        _set_progress(progress, 100, "Finished")
        QMessageBox.information(parent, "Success", f"Broadsheet exported to {path}")
    except Exception as e:
        print(f"[ERROR] Broadsheet export failed: {e}")
        QMessageBox.critical(parent, "Export Error", "An unexpected error occurred during export.")
    finally:
        if progress is not None:
            progress.close()
