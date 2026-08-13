import logging
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image as ExcelImage
from PySide6.QtCore import Qt
from PySide6.QtWidgets import QFileDialog, QMessageBox, QProgressDialog, QApplication
from db_utils import fetch_one
import os

logger = logging.getLogger(__name__)


def make_progress(parent, title):
    progress = QProgressDialog(title, None, 0, 100, parent)
    progress.setWindowTitle(title)
    progress.setWindowModality(Qt.WindowModal if parent else Qt.NonModal)
    progress.setAutoClose(False)
    progress.setAutoReset(False)
    progress.setMinimumDuration(0)
    progress.setValue(0)
    progress.setLabelText(f"{title}\n\nProgress: 0%")
    QApplication.processEvents()
    return progress


def set_progress(progress, percent, message):
    if progress is None:
        return
    progress.setValue(max(0, min(100, int(percent))))
    progress.setLabelText(f"{message}\n\nProgress: {max(0, min(100, int(percent)))}%")
    QApplication.processEvents()

def download_template(parent, filename, title, headers, instructions=None, samples=None):
    path, _ = QFileDialog.getSaveFileName(parent, "Download Template", filename, "Excel Files (*.xlsx)")
    if not path: return
    
    progress = make_progress(parent, "Creating template...")
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Template"
        set_progress(progress, 10, "Preparing template")

        # 1. School Header
        profile = fetch_one("""
            SELECT school_name, school_address, school_phone, school_email, school_logo
            FROM school_profile
            LIMIT 1
        """)

        school_name = profile[0].upper() if profile and profile[0] else "SCHOOL MANAGEMENT SYSTEM"
        school_contact = f"{profile[1] if profile and profile[1] else '-'} | {profile[2] if profile and profile[2] else '-'} | {profile[3] if profile and profile[3] else '-'}"
        school_logo = profile[4] if profile and len(profile) > 4 else None

        # Styling
        blue_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        
        # Row 1-3: School Info & Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.cell(row=1, column=1, value=school_name).font = Font(size=16, bold=True)
        ws.cell(row=1, column=1).alignment = center_align

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        ws.cell(row=2, column=1, value=school_contact).font = Font(size=10)
        ws.cell(row=2, column=1).alignment = center_align

        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(headers))
        ws.cell(row=3, column=1, value=title.upper()).font = Font(size=14, bold=True, color="2563EB")
        ws.cell(row=3, column=1).alignment = center_align

        if school_logo and os.path.exists(school_logo):
            try:
                logo = ExcelImage(school_logo)
                logo.width = 72
                logo.height = 72
                ws.add_image(logo, "A1")
            except Exception:
                pass
        set_progress(progress, 35, "Adding school header")

        # Row 5: Instructions
        ws.cell(row=5, column=1, value="INSTRUCTIONS:").font = bold_font
        if not instructions:
            instructions = [
                "1. Do not modify the column headers below.",
                "2. Start data entry from the row below the sample row.",
                "3. Ensure the 'Admission No' exists in the system where required.",
                "4. Keep the sample row as a formatting guide and replace it with real data."
            ]
        instructions_start_row = 6
        for i, text in enumerate(instructions):
            ws.cell(row=instructions_start_row + i, column=1, value=text).font = Font(italic=True, size=9)
        set_progress(progress, 55, "Writing instructions")

        # Headers row: placed dynamically after the instructions, with a blank spacer row
        header_row = instructions_start_row + len(instructions) + 1
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.fill = blue_fill
            cell.font = white_font
            cell.alignment = center_align
        set_progress(progress, 75, "Writing headers")

        # Sample row: directly below the headers
        sample_row = header_row + 1
        if samples:
            for col_num, val in enumerate(samples, 1):
                cell = ws.cell(row=sample_row, column=col_num, value=val)
                cell.font = Font(italic=True, color="808080")
        set_progress(progress, 90, "Adding sample row")

        # Freeze Panes (Header stays visible)
        ws.freeze_panes = f"A{sample_row}"

        # Auto Column Width — only measure from the header row downward so the
        # school-name/instructions block (all in column A) doesn't blow out col A's width
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for row_num in range(header_row, ws.max_row + 1):
                cell = ws.cell(row=row_num, column=col_num)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max_length + 4

        wb.save(path)
        set_progress(progress, 100, "Template saved")
        QMessageBox.information(parent, "Success", f"Template saved to {path}")
    except Exception as e:
        logger.exception("Failed to generate enrollment template")
        QMessageBox.critical(parent, "Error", "Failed to save template. Please check the file path and try again.")
    finally:
        progress.close()

def export_to_excel(parent, filename, headers, data, title=None):
    path, _ = QFileDialog.getSaveFileName(parent, "Export Data", filename, "Excel Files (*.xlsx)")
    if not path: return
    
    progress = make_progress(parent, "Exporting data...")
    try:
        wb = openpyxl.Workbook()
        ws = wb.active

        blue_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        ncols = len(headers)

        # School branding block, matching download_template()'s layout.
        # No "INSTRUCTIONS:" label is written here — exports aren't meant
        # to be re-imported, and omitting that label keeps find_header_row()
        # correctly rejecting an exported file if someone tries anyway,
        # rather than silently misreading the branding rows as data.
        set_progress(progress, 10, "Preparing export")
        profile = fetch_one("""
            SELECT school_name, school_address, school_phone, school_email, school_logo
            FROM school_profile
            LIMIT 1
        """)
        school_name = profile[0].upper() if profile and profile[0] else "SCHOOL MANAGEMENT SYSTEM"
        school_contact = f"{profile[1] if profile and profile[1] else '-'} | {profile[2] if profile and profile[2] else '-'} | {profile[3] if profile and profile[3] else '-'}"
        school_logo = profile[4] if profile and len(profile) > 4 else None

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        ws.cell(row=1, column=1, value=school_name).font = Font(size=16, bold=True)
        ws.cell(row=1, column=1).alignment = center_align

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        ws.cell(row=2, column=1, value=school_contact).font = Font(size=10)
        ws.cell(row=2, column=1).alignment = center_align

        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
        ws.cell(row=3, column=1, value=(title or "DATA EXPORT").upper()).font = Font(size=14, bold=True, color="2563EB")
        ws.cell(row=3, column=1).alignment = center_align

        if school_logo and os.path.exists(school_logo):
            try:
                logo = ExcelImage(school_logo)
                logo.width = 72
                logo.height = 72
                ws.add_image(logo, "A1")
            except Exception:
                pass

        header_row = 5
        # Header row: same blue-fill/bold-white convention as download_template.
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.fill = blue_fill
            cell.font = white_font
            cell.alignment = center_align

        data_start_row = header_row + 1
        total = max(len(data), 1)
        for index, row in enumerate(data, start=0):
            row_num = data_start_row + index
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = center_align
            if index == 0 or index == total - 1 or index % 10 == 0:
                set_progress(progress, 10 + int(((index + 1) / total) * 80), "Writing rows")

        # Auto column width from actual content (header + data), so long
        # names/admission numbers aren't clipped and short tick-mark columns
        # don't sit unnecessarily wide.
        for col_num in range(1, ncols + 1):
            column_letter = get_column_letter(col_num)
            max_length = len(str(headers[col_num - 1]))
            for row in data:
                if col_num - 1 < len(row) and row[col_num - 1]:
                    max_length = max(max_length, len(str(row[col_num - 1])))
            ws.column_dimensions[column_letter].width = max_length + 4

        ws.row_dimensions[header_row].height = 22
        ws.freeze_panes = f"A{data_start_row}"
        wb.save(path)
        set_progress(progress, 100, "Export complete")
        QMessageBox.information(parent, "Success", f"Data exported to {path}")
    except Exception as e:
        logger.exception("Failed to export enrollment data")
        QMessageBox.critical(parent, "Error", "Failed to export data. Please check the file path and try again.")
    finally:
        progress.close()

def get_import_file(parent):
    path, _ = QFileDialog.getOpenFileName(parent, "Select Excel File", "", "Excel Files (*.xlsx *.xls)")
    return path


def find_header_row(ws, instructions_label="INSTRUCTIONS:", max_scan_row=60):
    """Locate the header row written by download_template().

    The generated layout is always: a row with `instructions_label` in
    column A, one row per instruction line (also column A, no gaps), a
    single blank spacer row, then the header row. Scanning for this
    structure (rather than hardcoding a row number) keeps import code
    correct regardless of how many instructions a given template call
    passed in — the header/sample/data rows all shift depending on that
    count.

    Returns the 1-indexed header row number, or None if the expected
    structure isn't found (e.g. a hand-edited or non-template file), so
    callers can surface a clear error instead of silently misreading data.
    """
    label_row = None
    for row in range(1, max_scan_row + 1):
        value = ws.cell(row=row, column=1).value
        if value and str(value).strip().upper() == instructions_label:
            label_row = row
            break
    if label_row is None:
        return None

    row = label_row + 1
    while row <= max_scan_row and ws.cell(row=row, column=1).value not in (None, ""):
        row += 1
    # `row` is now the first blank spacer row after the instructions; the
    # header row immediately follows it.
    header_row = row + 1
    if header_row > max_scan_row:
        return None
    return header_row


def find_data_start_row(ws, instructions_label="INSTRUCTIONS:", max_scan_row=60):
    """Header row + 1 (sample row) + 1 = first real data row. Returns None
    if the header row itself can't be located."""
    header_row = find_header_row(ws, instructions_label=instructions_label, max_scan_row=max_scan_row)
    if header_row is None:
        return None
    return header_row + 2